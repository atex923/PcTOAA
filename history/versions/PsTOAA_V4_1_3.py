# BudgetAnalyzer_QT_V4_1_3.py
# 預算書分析系統 QT V4.1.3
#
# PySide6 / PyQt6 版
# - QTableWidget 支援單一儲存格字色
# - 表格顯示格線
# - 保留核心功能：
#   1. 讀取預算總表 / 預算詳細表 / 預算單價分析表
#   2. 組合總表
#   3. 輸出總表
#   4. 刪除單複價
#   5. 計算工項複價 / 重新計算
#   6. 整理備註
#   7. 匯出 XLS / XLSM
#   8. 匯入回饋 Excel
#   9. 輸出總表動態填入檔別、動支單號、單位檢核末碼
#   10. 單一儲存格顏色：
#      - 紅色：刪除規則後保留的單價/複價
#      - 紫色：手動修改
#      - 藍色：重新計算產生的新數字
# - 修正手動修改欄位資料功能，可即時切換可編輯狀態
# - 重新整理上方按鈕與欄框排列
# - 手動修改只標示該儲存格紫色，並依項目同步其他分頁同欄位資料
# - V3.6.1 修正：
#   勾選「啟用資料編輯」時，原本會逐格 setFlags，
#   可能連續觸發 itemChanged，進而同步染色與跨分頁同步，
#   導致大量遞迴/事件風暴而死當。
#   本版加入內部更新鎖與表格訊號暫停，避免啟用編輯時卡死。
# - V4.0.C1 新增：
#   1. 儲存編輯紀錄到 .budget_log.json，關閉後可重新載入繼續編輯。
#   2. 手動儲存 / 載入編輯紀錄按鈕。
#   3. 最多保留 10 次編輯歷史，可回復上一次編輯。
#   4. 上半部右側動態顯示 01發包工程費 / 011包工程 / 0B營業稅雙欄 / 總經費，米黃色底、紅字三位小數。
# - V4.1.1 精簡優化：
#   1. 合併重複解析邏輯與表格設定。
#   2. 修正靜默更新時 signals 還原狀態。
#   3. 優化組合總表父子項目配對。
# - V4.1.2 名稱調整：
#   1. 分頁顯示名稱統一為組合總表 / 輸出總表。
#   2. 功能按鈕名稱統一為刪除單複價 / 整理備註。
# - V4.1.3 名稱修正：
#   1. 組和總表修正為組合總表。

import sys
import re
import html
import traceback
import subprocess
import importlib
import json
import copy
from datetime import datetime
from pathlib import Path
from html.parser import HTMLParser


# =========================================================
# 缺少套件自動安裝
# =========================================================
def ensure_package(import_name, pip_name=None):
    """
    嘗試匯入套件；若缺少，使用目前 Python 執行環境自動 pip install。
    """
    pip_name = pip_name or import_name

    try:
        return importlib.import_module(import_name)
    except ImportError:
        print(f"[套件缺少] {import_name}，開始安裝：{pip_name}")

        try:
            subprocess.check_call([
                sys.executable,
                "-m",
                "pip",
                "install",
                pip_name
            ])
        except Exception as install_error:
            print(f"[安裝失敗] {pip_name}")
            raise install_error

        return importlib.import_module(import_name)


# Excel / 資料套件
pd = ensure_package("pandas", "pandas")

ensure_package("openpyxl", "openpyxl")
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# .xls 讀取套件，pandas 讀取 .xls 時會用到
ensure_package("xlrd", "xlrd")


# GUI 套件：優先使用 PySide6，若安裝失敗再嘗試 PyQt6
try:
    ensure_package("PySide6", "PySide6")

    from PySide6.QtCore import Qt
    from PySide6.QtGui import QColor, QBrush
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
        QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit,
        QPushButton, QCheckBox, QTabWidget, QTableWidget, QTableWidgetItem,
        QHeaderView, QInputDialog, QAbstractItemView
    )

except Exception:
    ensure_package("PyQt6", "PyQt6")

    from PyQt6.QtCore import Qt
    from PyQt6.QtGui import QColor, QBrush
    from PyQt6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
        QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit,
        QPushButton, QCheckBox, QTabWidget, QTableWidget, QTableWidgetItem,
        QHeaderView, QInputDialog, QAbstractItemView
    )


CODES = "123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
APP_VERSION = "V4.1.3"
APP_TITLE = f"預算書分析系統 QT {APP_VERSION}"
DEFAULT_LOG_FILENAME = "BudgetAnalyzer_V4_1_3.budget_log.json"

COLUMNS = (
    "檔別", "動支單號", "項目", "項目名稱", "來源編號", "來源名稱",
    "序列", "材料編號", "單位", "數量", "單價", "複價",
    "比例", "計算式", "規格", "備註", "單位檢核末碼"
)

AMOUNT_COLUMNS = {"數量", "單價", "複價"}
LEFT_COLUMNS = {"項目", "項目名稱", "備註"}
EMPTY_SHRINK_COLUMNS = {
    "來源編號", "來源名稱", "序列", "材料編號",
    "比例", "計算式", "規格"
}


class SimpleHTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self.current_row = None
        self.current_cell = None
        self.in_cell = False

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "tr":
            self.current_row = []
        elif tag in ("td", "th"):
            self.current_cell = []
            self.in_cell = True

    def handle_data(self, data):
        if self.in_cell and self.current_cell is not None:
            self.current_cell.append(data)

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("td", "th"):
            value = "".join(self.current_cell or [])
            value = html.unescape(value).strip()
            if self.current_row is not None:
                self.current_row.append(value)
            self.current_cell = None
            self.in_cell = False
        elif tag == "tr":
            if self.current_row is not None:
                self.rows.append(self.current_row)
            self.current_row = None


class BudgetAnalyzerQT(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle(APP_TITLE)
        self.resize(1900, 1000)

        self.data = []
        self.detail_data = []
        self.unit_price_data = []
        self.combined_data = []
        self.final_data = []

        self.color_marks = {}  # (row, col_name) -> QColor，組合總表計算/刪除顏色
        self.edited_cell_marks = {}  # (table_name, row, col_name) -> QColor，手動修改紫色

        self.is_loading_table = False
        self.is_internal_change = False
        self.edit_history = []
        self.max_history = 10

        self.build_ui()

    # =========================================================
    # 共用工具
    # =========================================================
    def table_configs(self):
        return (
            ("summary", self.data, self.summary_table),
            ("detail", self.detail_data, self.detail_table),
            ("unit_price", self.unit_price_data, self.unit_price_table),
            ("combined", self.combined_data, self.combined_table),
            ("final", self.final_data, self.final_table),
        )

    def all_tables(self):
        return tuple(table for _, _, table in self.table_configs())

    def msg_info(self, title, text):
        QMessageBox.information(self, title, text)

    def msg_warn(self, title, text):
        QMessageBox.warning(self, title, text)

    def msg_error(self, title, text):
        QMessageBox.critical(self, title, text)

    def fmt(self, value):
        try:
            if pd.isna(value) or value == "":
                return ""
            return f"{float(value):,.2f}"
        except Exception:
            return "" if value is None else str(value)

    def fmt_3(self, value):
        try:
            if pd.isna(value) or value == "":
                return ""
            return f"{float(value):,.3f}"
        except Exception:
            return "" if value is None else str(value)

    def num(self, value):
        try:
            text = "" if value is None else str(value).replace(",", "").strip()
            return None if text == "" else float(text)
        except Exception:
            return None

    def text(self, value):
        return "" if pd.isna(value) else str(value).strip()

    def compact(self, value):
        return "" if pd.isna(value) else "".join(str(value).strip().split())

    def is_blank(self, value):
        return self.fmt(value) == ""

    def engine_for(self, filename):
        return "xlrd" if filename.lower().endswith(".xls") else "openpyxl"

    def record(self, item="", name="", unit="", qty="", unit_price="", amount="", remark="", price3=False):
        return {
            "檔別": self.file_type_edit.text(),
            "動支單號": self.payment_no_edit.text(),
            "項目": item,
            "項目名稱": name,
            "來源編號": "",
            "來源名稱": "",
            "序列": "",
            "材料編號": "",
            "單位": unit,
            "數量": self.fmt(qty),
            "單價": self.fmt_3(unit_price) if price3 else self.fmt(unit_price),
            "複價": self.fmt_3(amount) if price3 else self.fmt(amount),
            "比例": "",
            "計算式": "",
            "規格": "",
            "備註": remark,
            "單位檢核末碼": self.unit_check_edit.text(),
        }

    # =========================================================
    # UI
    # =========================================================
    def build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)

        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(8, 6, 8, 6)
        main_layout.setSpacing(5)

        # =====================================================
        # V4.1.3：上半部緊密版面 + 右上雙欄費用比對
        # 左側功能區 + 右側米黃色金額資訊區，減少左右空白落差。
        # =====================================================
        top = QWidget()
        top_layout = QHBoxLayout(top)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(8)
        main_layout.addWidget(top)

        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(3)
        top_layout.addWidget(left_panel, stretch=1)

        def compact_button(text, slot, min_width=88):
            button = QPushButton(text)
            button.clicked.connect(slot)
            button.setMinimumWidth(min_width)
            button.setFixedHeight(30)
            button.setStyleSheet("padding: 2px 8px;")
            return button

        def small_label(text):
            label = QLabel(text)
            label.setStyleSheet("font-weight: bold;")
            return label

        self.file_edit = QLineEdit()
        self.project_name_edit = QLineEdit()
        self.project_no_edit = QLineEdit()
        self.file_type_edit = QLineEdit()
        self.payment_no_edit = QLineEdit()
        self.unit_check_edit = QLineEdit()
        self.hide_level_edit = QLineEdit()

        self.file_edit.setMinimumWidth(420)
        self.project_name_edit.setMinimumWidth(360)
        self.project_no_edit.setFixedWidth(145)
        self.file_type_edit.setFixedWidth(80)
        self.payment_no_edit.setFixedWidth(135)
        self.unit_check_edit.setFixedWidth(75)
        self.hide_level_edit.setFixedWidth(52)

        for edit in (
            self.file_edit, self.project_name_edit, self.project_no_edit,
            self.file_type_edit, self.payment_no_edit, self.unit_check_edit,
            self.hide_level_edit,
        ):
            edit.setFixedHeight(28)

        # -----------------------------------------------------
        # 第1列：檔案選擇與主要操作
        # -----------------------------------------------------
        file_row = QHBoxLayout()
        file_row.setSpacing(4)
        file_row.addWidget(small_label("Excel檔案"))
        file_row.addWidget(self.file_edit, stretch=1)
        file_row.addWidget(compact_button("瀏覽", self.browse_file, 64))
        file_row.addWidget(compact_button("分析", self.analyze, 64))
        file_row.addWidget(compact_button("自動欄寬", self.auto_resize_all, 86))
        left_layout.addLayout(file_row)

        # -----------------------------------------------------
        # 第2列：工程資訊
        # -----------------------------------------------------
        project_row = QHBoxLayout()
        project_row.setSpacing(4)
        project_row.addWidget(small_label("工程名稱"))
        project_row.addWidget(self.project_name_edit, stretch=1)
        project_row.addWidget(small_label("工程編號"))
        project_row.addWidget(self.project_no_edit)
        left_layout.addLayout(project_row)

        # -----------------------------------------------------
        # 第3列：固定欄位與編輯設定
        # -----------------------------------------------------
        fixed_row = QHBoxLayout()
        fixed_row.setSpacing(4)
        fixed_row.addWidget(small_label("檔別"))
        fixed_row.addWidget(self.file_type_edit)
        fixed_row.addWidget(small_label("動支單號"))
        fixed_row.addWidget(self.payment_no_edit)
        fixed_row.addWidget(small_label("單位檢核末碼"))
        fixed_row.addWidget(self.unit_check_edit)

        self.edit_check = QCheckBox("啟用資料編輯")
        self.edit_check.toggled.connect(self.update_editable_state)
        self.manual_width_check = QCheckBox("手動調整欄寬")
        fixed_row.addSpacing(8)
        fixed_row.addWidget(self.edit_check)
        fixed_row.addWidget(self.manual_width_check)
        fixed_row.addStretch(1)
        left_layout.addLayout(fixed_row)

        # -----------------------------------------------------
        # 第4列：組合總表分析與顯示控制
        # -----------------------------------------------------
        analysis_row = QHBoxLayout()
        analysis_row.setSpacing(4)
        analysis_row.addWidget(small_label("組合總表"))
        analysis_row.addWidget(compact_button("刪除單複價", self.apply_delete_rules, 92))
        analysis_row.addWidget(compact_button("計算工項複價", self.calculate_leaf_amounts, 104))
        analysis_row.addWidget(compact_button("重新計算", self.calculate_rollup_amounts, 86))
        analysis_row.addWidget(small_label("隱藏階層"))
        analysis_row.addWidget(self.hide_level_edit)
        analysis_row.addWidget(compact_button("隱藏", self.hide_combined_below_level, 58))
        analysis_row.addWidget(compact_button("恢復", self.show_all_combined, 58))
        analysis_row.addWidget(compact_button("整理備註", self.clean_combined_remarks, 82))
        analysis_row.addStretch(1)
        left_layout.addLayout(analysis_row)

        # -----------------------------------------------------
        # 第5列：輸出總表與檔案輸出入 / 編輯紀錄
        # -----------------------------------------------------
        io_row = QHBoxLayout()
        io_row.setSpacing(4)
        io_row.addWidget(small_label("輸出總表"))
        io_row.addWidget(compact_button("填入欄位", self.fill_final_fixed_fields, 82))
        io_row.addWidget(compact_button("儲存編輯", self.save_edit_log_manual, 82))
        io_row.addWidget(compact_button("載入編輯", self.load_edit_log_manual, 82))
        io_row.addWidget(compact_button("回復一次", self.undo_last_edit, 82))
        io_row.addWidget(compact_button("匯出 XLS", self.export_final_xls, 82))
        io_row.addWidget(compact_button("匯出 XLSM", self.export_final_xlsm, 88))
        io_row.addWidget(compact_button("匯入回饋", self.import_feedback_excel, 82))
        io_row.addStretch(1)
        left_layout.addLayout(io_row)

        # -----------------------------------------------------
        # 右側：動態金額資訊區
        # -----------------------------------------------------
        self.amount_panel = QWidget()
        self.amount_panel.setFixedWidth(720)
        self.amount_panel.setStyleSheet(
            "background-color: #FFF2CC; "
            "border: 1px solid #D6B656; "
            "border-radius: 8px;"
        )
        amount_panel_layout = QGridLayout(self.amount_panel)
        amount_panel_layout.setContentsMargins(12, 6, 12, 6)
        amount_panel_layout.setHorizontalSpacing(8)
        amount_panel_layout.setVerticalSpacing(2)

        # V4.1.3：右上資訊區字體縮小約 20%，避免雙欄資料過擠。
        name_style = "color: red; font-weight: bold; font-size: 13px; background-color: transparent;"
        value_style = "color: red; font-weight: bold; font-size: 13px; background-color: transparent;"
        calc_name_style = "color: red; font-weight: bold; font-size: 12px; background-color: transparent;"

        self.amount_01_name_label = QLabel("01發包工程費")
        self.amount_011_name_label = QLabel("011包工程")
        self.amount_0b_name_label = QLabel("0B營業稅")
        self.amount_profit_name_label = QLabel("承包商利潤及工程保險費")
        self.amount_air_name_label = QLabel("空氣污染防制費")
        self.amount_qc_name_label = QLabel("二級品管抽驗費")
        self.amount_total_name_label = QLabel("總經費")

        self.amount_0b_calc_name_label = QLabel("01×5%")
        self.amount_profit_calc_name_label = QLabel("01×10%")
        self.amount_air_calc_name_label = QLabel("01×0.5%")
        self.amount_qc_calc_name_label = QLabel("01×0.1%")

        self.amount_01_value_label = QLabel("0.000")
        self.amount_011_value_label = QLabel("0.000")
        self.amount_0b_value_label = QLabel("0.000")
        self.amount_0b_calc_value_label = QLabel("0.000")
        self.amount_profit_value_label = QLabel("0.000")
        self.amount_profit_calc_value_label = QLabel("0.000")
        self.amount_air_value_label = QLabel("0.000")
        self.amount_air_calc_value_label = QLabel("0.000")
        self.amount_qc_value_label = QLabel("0.000")
        self.amount_qc_calc_value_label = QLabel("0.000")
        self.amount_total_value_label = QLabel("0.000")

        name_labels = [
            self.amount_01_name_label,
            self.amount_011_name_label,
            self.amount_0b_name_label,
            self.amount_profit_name_label,
            self.amount_air_name_label,
            self.amount_qc_name_label,
            self.amount_total_name_label,
        ]
        calc_name_labels = [
            self.amount_0b_calc_name_label,
            self.amount_profit_calc_name_label,
            self.amount_air_calc_name_label,
            self.amount_qc_calc_name_label,
        ]
        value_labels = [
            self.amount_01_value_label,
            self.amount_011_value_label,
            self.amount_0b_value_label,
            self.amount_0b_calc_value_label,
            self.amount_profit_value_label,
            self.amount_profit_calc_value_label,
            self.amount_air_value_label,
            self.amount_air_calc_value_label,
            self.amount_qc_value_label,
            self.amount_qc_calc_value_label,
            self.amount_total_value_label,
        ]

        for label in name_labels:
            label.setStyleSheet(name_style)
            label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        for label in calc_name_labels:
            label.setStyleSheet(calc_name_style)
            label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        for label in value_labels:
            label.setStyleSheet(value_style)
            label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self.amount_panel.setMinimumHeight(176)
        for label in (
            self.amount_01_value_label, self.amount_011_value_label, self.amount_total_value_label
        ):
            label.setMinimumWidth(180)
        for label in (
            self.amount_0b_value_label, self.amount_0b_calc_value_label,
            self.amount_profit_value_label, self.amount_profit_calc_value_label,
            self.amount_air_value_label, self.amount_air_calc_value_label,
            self.amount_qc_value_label, self.amount_qc_calc_value_label,
        ):
            label.setMinimumWidth(120)

        # 一般金額列：左側項目名稱，右側金額。
        amount_panel_layout.addWidget(self.amount_01_name_label, 0, 0)
        amount_panel_layout.addWidget(self.amount_01_value_label, 0, 1, 1, 3)
        amount_panel_layout.addWidget(self.amount_011_name_label, 1, 0)
        amount_panel_layout.addWidget(self.amount_011_value_label, 1, 1, 1, 3)

        # 雙欄列：左邊抓組合總表複價，右邊依 01 發包工程費複價試算。
        dual_rows = [
            (2, self.amount_0b_name_label, self.amount_0b_value_label, self.amount_0b_calc_name_label, self.amount_0b_calc_value_label),
            (3, self.amount_profit_name_label, self.amount_profit_value_label, self.amount_profit_calc_name_label, self.amount_profit_calc_value_label),
            (4, self.amount_air_name_label, self.amount_air_value_label, self.amount_air_calc_name_label, self.amount_air_calc_value_label),
            (5, self.amount_qc_name_label, self.amount_qc_value_label, self.amount_qc_calc_name_label, self.amount_qc_calc_value_label),
        ]

        for row_no, title_label, actual_label, calc_label, calc_value_label in dual_rows:
            amount_panel_layout.addWidget(title_label, row_no, 0)
            amount_panel_layout.addWidget(actual_label, row_no, 1)
            amount_panel_layout.addWidget(calc_label, row_no, 2)
            amount_panel_layout.addWidget(calc_value_label, row_no, 3)

        amount_panel_layout.addWidget(self.amount_total_name_label, 6, 0)
        amount_panel_layout.addWidget(self.amount_total_value_label, 6, 1, 1, 3)

        self.amount_total_name_label.setStyleSheet(
            name_style + "border-top: 1px solid #D6B656; padding-top: 3px;"
        )
        self.amount_total_value_label.setStyleSheet(
            value_style + "border-top: 1px solid #D6B656; padding-top: 3px;"
        )

        top_layout.addWidget(self.amount_panel, stretch=0)

        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs, stretch=1)

        self.summary_table = self.create_table()
        self.detail_table = self.create_table()
        self.unit_price_table = self.create_table()
        self.combined_table = self.create_table()
        self.final_table = self.create_table()

        self.tabs.addTab(self.summary_table, "預算總表")
        self.tabs.addTab(self.detail_table, "預算詳細表")
        self.tabs.addTab(self.unit_price_table, "預算單價分析表")
        self.tabs.addTab(self.combined_table, "組合總表")
        self.tabs.addTab(self.final_table, "輸出總表")

        self.status_label = QLabel("尚未分析")
        main_layout.addWidget(self.status_label)

    def create_table(self):
        table = QTableWidget()
        table.setColumnCount(len(COLUMNS))
        table.setHorizontalHeaderLabels(COLUMNS)
        table.setShowGrid(True)
        table.setGridStyle(Qt.PenStyle.SolidLine)
        table.setAlternatingRowColors(False)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.verticalHeader().setVisible(True)
        table.itemChanged.connect(self.on_item_changed)
        return table

    def table_data_ref(self, table):
        for name, data, target_table in self.table_configs():
            if table is target_table:
                return name, data
        return "", []

    def populate_table(self, table, data):
        self.is_loading_table = True
        table.setRowCount(0)
        table.setRowCount(len(data))

        for r, row in enumerate(data):
            for c, col in enumerate(COLUMNS):
                value = "" if row.get(col, "") is None else str(row.get(col, ""))
                item = QTableWidgetItem(value)

                if col in LEFT_COLUMNS:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                elif col in AMOUNT_COLUMNS:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                self.set_item_editable_flag(item)

                table.setItem(r, c, item)

        self.is_loading_table = False

        self.apply_manual_edit_marks_to_table(table)

        if table is self.combined_table:
            self.apply_color_marks_to_combined()
            self.update_dynamic_amount_labels()

        if not self.manual_width_check.isChecked():
            self.auto_resize_table(table)

        self.update_editable_state()

    def auto_resize_table(self, table):
        for c, col in enumerate(COLUMNS):
            if table is self.combined_table and col in EMPTY_SHRINK_COLUMNS:
                is_empty = True
                for r in range(table.rowCount()):
                    item = table.item(r, c)
                    if item and item.text().strip():
                        is_empty = False
                        break
                if is_empty:
                    table.setColumnWidth(c, 1)
                    continue

            table.resizeColumnToContents(c)
            width = table.columnWidth(c)
            if col in LEFT_COLUMNS:
                table.setColumnWidth(c, min(max(width + 20, 80), 700))
            else:
                table.setColumnWidth(c, min(max(width + 20, 70), 260))

    def auto_resize_all(self):
        for table in self.all_tables():
            self.auto_resize_table(table)
        self.msg_info("完成", "已自動調整欄寬")

    def update_editable_state(self):
        """
        修正啟用資料編輯會死當：
        切換可編輯狀態時需要逐格 setFlags。Qt 在某些環境會把 setFlags /
        setForeground 等內部變更也視為 itemChanged，造成 on_item_changed 被大量觸發，
        接著又進行跨分頁同步與染色，形成事件風暴。

        解法：
        1. 進入內部更新鎖 self.is_internal_change。
        2. 每張表格暫停 signals。
        3. 完成 flags 更新後恢復 signals。
        """
        editable = self.edit_check.isChecked()

        self.is_internal_change = True

        try:
            for table in self.all_tables():
                old_block_state = table.blockSignals(True)

                try:
                    if editable:
                        table.setEditTriggers(
                            QAbstractItemView.EditTrigger.DoubleClicked
                            | QAbstractItemView.EditTrigger.EditKeyPressed
                            | QAbstractItemView.EditTrigger.AnyKeyPressed
                        )
                    else:
                        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

                    for r in range(table.rowCount()):
                        for c in range(table.columnCount()):
                            item = table.item(r, c)

                            if item is None:
                                continue

                            flags = item.flags()

                            if editable:
                                item.setFlags(flags | Qt.ItemFlag.ItemIsEditable)
                            else:
                                item.setFlags(flags & ~Qt.ItemFlag.ItemIsEditable)

                finally:
                    table.blockSignals(old_block_state)

        finally:
            self.is_internal_change = False

    def set_item_editable_flag(self, item):
        """
        建立儲存格時依目前勾選狀態設定是否可編輯。
        """
        if self.edit_check.isChecked():
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        else:
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

    def format_dynamic_amount(self, value):
        """
        上半部金額顯示格式：#,##0.000。
        若資料空白或不是數字，顯示 0.000。
        """
        number = self.num(value)
        if number is None:
            number = 0.0
        return f"{number:,.3f}"

    def get_combined_amount_by_item(self, item_code):
        for row in self.combined_data:
            if str(row.get("項目", "")).strip().upper() == str(item_code).strip().upper():
                return row.get("複價", "")
        return ""

    def get_combined_amount_by_name(self, item_name):
        """
        依組合總表「項目名稱」抓複價。
        先做完全相同比對；找不到時再用包含比對，提高不同預算書命名的容錯。
        """
        target = self.compact(item_name)

        for row in self.combined_data:
            name = self.compact(row.get("項目名稱", ""))
            if name == target:
                return row.get("複價", "")

        for row in self.combined_data:
            name = self.compact(row.get("項目名稱", ""))
            if target and (target in name or name in target):
                return row.get("複價", "")

        return ""

    def update_dynamic_amount_labels(self):
        """
        動態抓取組合總表指定資料：
        - 01 發包工程費：項目 01 複價
        - 011 包工程：項目 011 複價
        - 0B 營業稅：左欄項目 0B 複價，右欄 01 × 5%
        - 承包商利潤及工程保險費：左欄項目 012 複價，右欄 01 × 10%
        - 空氣污染防制費：左欄項目 0A1 複價，右欄 01 × 0.5%
        - 二級品管抽驗費：左欄項目 0A2 複價，右欄 01 × 0.1%
        - 總經費 = 01 + 02 + 03 + 04 + 05 + 06 + 0A
        """
        required_labels = (
            "amount_01_value_label",
            "amount_011_value_label",
            "amount_0b_value_label",
            "amount_0b_calc_value_label",
            "amount_profit_value_label",
            "amount_profit_calc_value_label",
            "amount_air_value_label",
            "amount_air_calc_value_label",
            "amount_qc_value_label",
            "amount_qc_calc_value_label",
            "amount_total_value_label",
        )

        if not all(hasattr(self, name) for name in required_labels):
            return

        amounts = {
            code: self.num(self.get_combined_amount_by_item(code)) or 0.0
            for code in ("01", "011", "0B", "012", "0A1", "0A2", "02", "03", "04", "05", "06", "0A")
        }
        base_01 = amounts["01"]

        label_values = (
            (self.amount_01_value_label, amounts["01"]),
            (self.amount_011_value_label, amounts["011"]),
            (self.amount_0b_value_label, amounts["0B"]),
            (self.amount_0b_calc_value_label, base_01 * 0.05),
            (self.amount_profit_value_label, amounts["012"]),
            (self.amount_profit_calc_value_label, base_01 * 0.10),
            (self.amount_air_value_label, amounts["0A1"]),
            (self.amount_air_calc_value_label, base_01 * 0.005),
            (self.amount_qc_value_label, amounts["0A2"]),
            (self.amount_qc_calc_value_label, base_01 * 0.001),
            (self.amount_total_value_label, sum(amounts[code] for code in ("01", "02", "03", "04", "05", "06", "0A"))),
        )

        for label, value in label_values:
            label.setText(f"{value:,.3f}")

    def table_object_by_name(self, name):
        for table_name, _, table in self.table_configs():
            if table_name == name:
                return table
        return None

    def set_table_cell_value_silent(self, table, row, col_name, value):
        """
        靜默更新表格指定儲存格，避免觸發遞迴 itemChanged。
        """
        col_index = COLUMNS.index(col_name)
        old_block_state = table.blockSignals(True)

        try:
            item = table.item(row, col_index)

            if item is None:
                item = QTableWidgetItem("")
                table.setItem(row, col_index, item)

            item.setText("" if value is None else str(value))

        finally:
            table.blockSignals(old_block_state)

    def set_item_foreground_silent(self, table, item, color_name):
        if item is None:
            return

        old_block_state = table.blockSignals(True)
        old_internal_state = self.is_internal_change
        self.is_internal_change = True

        try:
            item.setForeground(QBrush(QColor(color_name)))
        finally:
            self.is_internal_change = old_internal_state
            table.blockSignals(old_block_state)

    def mark_manual_cell(self, table_name, row, col_name):
        """
        只標記手動修改的單一儲存格紫色。
        染色屬於程式內部視覺更新，必須暫停 signals，避免再次觸發 itemChanged。
        """
        self.edited_cell_marks[(table_name, row, col_name)] = "purple"

        table = self.table_object_by_name(table_name)

        if table is None:
            return

        col_index = COLUMNS.index(col_name)
        item = table.item(row, col_index)

        self.set_item_foreground_silent(table, item, "purple")

    def apply_manual_edit_marks_to_table(self, table):
        """
        表格重新載入後，恢復手動修改儲存格的紫色標記。
        """
        name, _ = self.table_data_ref(table)

        for (table_name, row, col_name), color in list(self.edited_cell_marks.items()):
            if table_name != name:
                continue

            if row >= table.rowCount():
                continue

            col_index = COLUMNS.index(col_name)
            item = table.item(row, col_index)

            if item:
                item.setForeground(QBrush(QColor(color)))

    def sync_edit_to_other_pages(self, source_name, source_row, item_code, col_name, value):
        """
        只要有一個分頁修改資料：
        依「項目」欄位，在其他分頁找到相同項目，更新同一欄位資料。
        組合總表/輸出總表互相同步；其他分頁也會同步相同項目。
        """
        if not item_code:
            return

        for target_name, target_data, target_table in self.table_configs():
            for idx, row in enumerate(target_data):
                if target_name == source_name and idx == source_row:
                    continue

                if str(row.get("項目", "")).strip() != item_code:
                    continue

                row[col_name] = value
                self.set_table_cell_value_silent(target_table, idx, col_name, value)

                # 同步更新的格子也標紫色，代表這格是因手動修改而變更。
                self.mark_manual_cell(target_name, idx, col_name)

    def on_item_changed(self, item):
        if self.is_loading_table or self.is_internal_change:
            return

        table = self.sender()

        if not isinstance(table, QTableWidget):
            return

        name, data = self.table_data_ref(table)

        if not name or not self.edit_check.isChecked():
            return

        r = item.row()
        c = item.column()
        col = COLUMNS[c]
        value = item.text()

        old_value = ""
        if 0 <= r < len(data):
            old_value = "" if data[r].get(col, "") is None else str(data[r].get(col, ""))
            if old_value == value:
                return
            self.push_edit_history()
            data[r][col] = value

        # 只讓被手動修改的這一格變成紫色。
        self.mark_manual_cell(name, r, col)

        # 依「項目」同步其他分頁相同項目的同欄位。
        item_code = ""
        if 0 <= r < len(data):
            item_code = str(data[r].get("項目", "")).strip()

        self.sync_edit_to_other_pages(name, r, item_code, col, value)
        self.update_dynamic_amount_labels()
        self.save_edit_log(show_message=False)


    # =========================================================
    # V4.1.3 儲存 / 載入 / 編輯歷史
    # =========================================================
    def log_path_for_current_file(self):
        """
        依目前 Excel 檔案產生專用 log 檔名。
        例：AAA.xlsx -> AAA.budget_log.json
        """
        filename = self.file_edit.text().strip()

        if filename:
            path = Path(filename)
            if path.parent.exists():
                return path.with_suffix(".budget_log.json")

        return Path.cwd() / DEFAULT_LOG_FILENAME

    def serialize_marks(self, marks):
        output = []

        for key, color in marks.items():
            output.append({"key": list(key), "color": str(color)})

        return output

    def deserialize_marks(self, rows):
        output = {}

        for row in rows or []:
            key = tuple(row.get("key", []))
            color = row.get("color", "")

            if key and color:
                output[key] = color

        return output

    def collect_state(self):
        return {
            "version": APP_VERSION,
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "file_path": self.file_edit.text(),
            "project_name": self.project_name_edit.text(),
            "project_no": self.project_no_edit.text(),
            "file_type": self.file_type_edit.text(),
            "payment_no": self.payment_no_edit.text(),
            "unit_check": self.unit_check_edit.text(),
            "hide_level": self.hide_level_edit.text(),
            "data": copy.deepcopy(self.data),
            "detail_data": copy.deepcopy(self.detail_data),
            "unit_price_data": copy.deepcopy(self.unit_price_data),
            "combined_data": copy.deepcopy(self.combined_data),
            "final_data": copy.deepcopy(self.final_data),
            "color_marks": self.serialize_marks(self.color_marks),
            "edited_cell_marks": self.serialize_marks(self.edited_cell_marks),
        }

    def apply_state(self, state):
        self.is_internal_change = True

        try:
            self.file_edit.setText(state.get("file_path", ""))
            self.project_name_edit.setText(state.get("project_name", ""))
            self.project_no_edit.setText(state.get("project_no", ""))
            self.file_type_edit.setText(state.get("file_type", ""))
            self.payment_no_edit.setText(state.get("payment_no", ""))
            self.unit_check_edit.setText(state.get("unit_check", ""))
            self.hide_level_edit.setText(state.get("hide_level", ""))

            self.data = copy.deepcopy(state.get("data", []))
            self.detail_data = copy.deepcopy(state.get("detail_data", []))
            self.unit_price_data = copy.deepcopy(state.get("unit_price_data", []))
            self.combined_data = copy.deepcopy(state.get("combined_data", []))
            self.final_data = copy.deepcopy(state.get("final_data", []))
            self.color_marks = self.deserialize_marks(state.get("color_marks", []))
            self.edited_cell_marks = self.deserialize_marks(state.get("edited_cell_marks", []))

            self.populate_table(self.summary_table, self.data)
            self.populate_table(self.detail_table, self.detail_data)
            self.populate_table(self.unit_price_table, self.unit_price_data)
            self.populate_table(self.combined_table, self.combined_data)
            self.populate_table(self.final_table, self.final_data)
            self.update_dynamic_amount_labels()

            self.status_label.setText(
                f"已載入編輯紀錄｜預算總表：{len(self.data)} 筆｜"
                f"預算詳細表：{len(self.detail_data)} 筆｜"
                f"單價分析：{len(self.unit_price_data)} 筆｜"
                f"組合表：{len(self.combined_data)} 筆｜"
                f"輸出總表：{len(self.final_data)} 筆"
            )

        finally:
            self.is_internal_change = False
            self.update_editable_state()

    def make_history_snapshot(self):
        state = self.collect_state()
        state["history_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return state

    def push_edit_history(self):
        self.edit_history.append(self.make_history_snapshot())

        if len(self.edit_history) > self.max_history:
            self.edit_history = self.edit_history[-self.max_history:]

    def save_edit_log(self, show_message=False):
        try:
            log_path = self.log_path_for_current_file()
            state = self.collect_state()
            payload = {
                "app": "BudgetAnalyzerQT",
                "version": APP_VERSION,
                "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "state": state,
                "history": self.edit_history[-self.max_history:],
            }

            log_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )

            self.status_label.setText(f"已儲存編輯紀錄：{log_path}")

            if show_message:
                self.msg_info("完成", f"已儲存編輯紀錄：\n{log_path}")

            return True

        except Exception:
            if show_message:
                self.msg_error("錯誤", traceback.format_exc())
            return False

    def save_edit_log_manual(self):
        self.save_edit_log(show_message=True)

    def load_edit_log_from_path(self, log_path):
        payload = json.loads(Path(log_path).read_text(encoding="utf-8"))
        state = payload.get("state", payload)
        self.edit_history = payload.get("history", [])[-self.max_history:]
        self.apply_state(state)

    def load_edit_log_manual(self):
        default_path = str(self.log_path_for_current_file())
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "載入編輯紀錄",
            default_path,
            "Budget Analyzer Log (*.budget_log.json *.json);;All Files (*.*)"
        )

        if not filename:
            return

        try:
            self.load_edit_log_from_path(filename)
            self.msg_info("完成", f"已載入編輯紀錄：\n{filename}")
        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    def ask_load_existing_log(self):
        log_path = self.log_path_for_current_file()

        if not log_path.exists():
            return

        reply = QMessageBox.question(
            self,
            "發現編輯紀錄",
            f"偵測到此檔案有編輯紀錄：\n{log_path}\n\n是否載入上次編輯內容？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.load_edit_log_from_path(log_path)
            except Exception:
                self.msg_error("錯誤", traceback.format_exc())

    def undo_last_edit(self):
        if not self.edit_history:
            self.msg_warn("提醒", "目前沒有可回復的編輯歷史。")
            return

        try:
            state = self.edit_history.pop()
            self.apply_state(state)
            self.save_edit_log(show_message=False)
            self.msg_info("完成", f"已回復上一次編輯。\n剩餘可回復次數：{len(self.edit_history)}")
        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    def closeEvent(self, event):
        self.save_edit_log(show_message=False)
        event.accept()

    # =========================================================
    # 讀取分析
    # =========================================================
    def browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "選擇預算書 Excel 檔案", "",
            "Excel (*.xlsx *.xls);;All Files (*.*)"
        )
        if path:
            self.file_edit.setText(path)
            self.ask_load_existing_log()

    def analyze(self):
        try:
            filename = self.file_edit.text().strip()

            if not filename:
                self.msg_warn("提醒", "請先選擇 Excel 檔案")
                return

            engine = self.engine_for(filename)

            self.parse_summary(filename, engine)
            self.parse_detail(filename, engine)
            warnings = self.parse_unit_price(filename, engine)
            self.build_combined_data()

            self.populate_table(self.summary_table, self.data)
            self.populate_table(self.detail_table, self.detail_data)
            self.populate_table(self.unit_price_table, self.unit_price_data)
            self.refresh_combined_and_final(self.combined_data)

            self.status_label.setText(
                f"分析完成｜預算總表：{len(self.data)} 筆｜"
                f"預算詳細表：{len(self.detail_data)} 筆｜"
                f"預算單價分析表：{len(self.unit_price_data)} 筆｜"
                f"組合表：{len(self.combined_data)} 筆｜"
                f"輸出總表：{len(self.final_data)} 筆"
            )

            self.edit_history = []
            self.save_edit_log(show_message=False)

            if warnings:
                self.msg_warn("單價分析材料超過 35 筆", "\n".join(warnings[:20]) + ("\n..." if len(warnings) > 20 else ""))
            else:
                self.msg_info("完成", "分析完成")

        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    def parse_budget_rows(self, df, target, has_unit_price_fields):
        target.clear()
        current = None

        for r in range(8, len(df)):
            item = self.text(df.iloc[r, 0])
            name = self.text(df.iloc[r, 1])
            remark = self.text(df.iloc[r, 6])

            if item and not item[0].isdigit():
                break

            if item:
                if current:
                    target.append(self.record(**current))

                current = {
                    "item": item,
                    "name": self.compact(name),
                    "unit": self.text(df.iloc[r, 2]) if has_unit_price_fields else "",
                    "qty": df.iloc[r, 3] if has_unit_price_fields else "",
                    "unit_price": df.iloc[r, 4] if has_unit_price_fields else "",
                    "amount": df.iloc[r, 5],
                    "remark": self.compact(remark),
                }

            elif current:
                current["name"] += self.compact(name)
                current["remark"] += self.compact(remark)

        if current:
            target.append(self.record(**current))

    def parse_summary(self, filename, engine):
        df = pd.read_excel(filename, sheet_name="預算總表", header=None, engine=engine)
        self.project_name_edit.setText(self.text(df.iloc[5, 1]))
        self.project_no_edit.setText(self.text(df.iloc[6, 5]))
        self.parse_budget_rows(df, self.data, False)

    def parse_detail(self, filename, engine):
        df = pd.read_excel(filename, sheet_name="預算詳細表", header=None, engine=engine)
        self.parse_budget_rows(df, self.detail_data, True)

    def parse_unit_price(self, filename, engine):
        df = pd.read_excel(filename, sheet_name="預算單價分析表", header=None, engine=engine)
        self.unit_price_data.clear()
        warnings = []

        r = 7

        while r < len(df):
            parent = self.text(df.iloc[r, 0])

            if not parent:
                r += 1
                continue

            if not parent.startswith("0"):
                break

            parent_unit = self.text(df.iloc[r, 2])

            material_count = 0
            current = None
            m = r + 2

            total_qty = None
            scan = m

            while scan < len(df):
                scan_name = self.text(df.iloc[scan, 1])
                if scan_name == "合計":
                    total_qty = self.num(df.iloc[scan, 3])
                    break
                scan += 1

            def calc_material_qty(raw_qty, raw_unit_price):
                if total_qty is None or total_qty == 0:
                    return raw_qty

                if parent_unit == "式":
                    base = self.num(raw_unit_price)
                else:
                    base = self.num(raw_qty)

                if base is None:
                    return raw_qty

                return base / total_qty

            def flush():
                nonlocal material_count, current

                if not current:
                    return

                material_count += 1

                if material_count > 35:
                    warnings.append(f"工項 {parent} 的單價分析材料超過 35 筆，第 {material_count} 筆起未匯入。")
                else:
                    current["item"] = f"{parent}{CODES[material_count - 1]}"
                    self.unit_price_data.append(self.record(**current, price3=True))

                current = None

            while m < len(df):
                a_value = self.text(df.iloc[m, 0])
                name = self.text(df.iloc[m, 1])
                unit = self.text(df.iloc[m, 2])
                raw_qty = df.iloc[m, 3]
                unit_price = df.iloc[m, 4]
                amount = df.iloc[m, 5]
                remark = self.text(df.iloc[m, 6])

                # 若讀到表頭列，略過：
                # C欄 = 單位 或 D欄 = 數量
                if unit == "單位" or self.text(raw_qty) == "數量":
                    m += 1
                    continue

                if name == "合計":
                    flush()
                    break

                blank_row = (
                    not a_value and not name and not unit
                    and self.is_blank(raw_qty) and self.is_blank(unit_price)
                    and self.is_blank(amount) and not remark
                )

                if blank_row:
                    m += 1
                    continue

                qty = calc_material_qty(raw_qty, unit_price)

                if current and not unit:
                    current["name"] += self.compact(name)
                    if remark:
                        current["remark"] += self.compact(remark)
                    if self.is_blank(current["qty"]) and not self.is_blank(qty):
                        current["qty"] = qty
                    if self.is_blank(current["unit_price"]) and not self.is_blank(unit_price):
                        current["unit_price"] = unit_price
                    if self.is_blank(current["amount"]) and not self.is_blank(amount):
                        current["amount"] = amount
                else:
                    flush()
                    current = {
                        "item": "",
                        "name": self.compact(name),
                        "unit": unit,
                        "qty": qty,
                        "unit_price": unit_price,
                        "amount": amount,
                        "remark": self.compact(remark),
                    }

                m += 1

            flush()
            r = m + 1

        return warnings

    # =========================================================
    # 組合總表 / 輸出總表
    # =========================================================
    def build_combined_data(self):
        self.combined_data.clear()

        detail_items = [
            str(row.get("項目", "")).strip()
            for row in self.detail_data
            if str(row.get("項目", "")).strip()
        ]

        child_map = {item: [] for item in detail_items}
        detail_item_set = set(detail_items)
        unmatched = []

        for row in self.unit_price_data:
            item = str(row.get("項目", "")).strip()
            parent = next(
                (item[:length] for length in range(len(item) - 1, 0, -1) if item[:length] in detail_item_set),
                ""
            )

            row = dict(row)
            row["來源名稱"] = ""

            if parent:
                child_map.setdefault(parent, []).append(row)
            else:
                unmatched.append(row)

        for row in self.detail_data:
            parent = str(row.get("項目", "")).strip()
            new_row = dict(row)
            new_row["來源名稱"] = ""
            self.combined_data.append(new_row)
            self.combined_data.extend(child_map.get(parent, []))

        for row in unmatched:
            row["來源名稱"] = ""
            self.combined_data.append(row)

    def final_number_text(self, value):
        text = "" if value is None else str(value).strip()
        if text == "":
            return ""
        return text.replace(",", "")

    def make_final_rows(self, source_data):
        final_rows = []

        for row in source_data:
            new_row = dict(row)
            new_row["檔別"] = self.file_type_edit.text()
            new_row["動支單號"] = self.payment_no_edit.text()
            new_row["單位檢核末碼"] = self.unit_check_edit.text()

            for col in ("數量", "單價", "複價"):
                new_row[col] = self.final_number_text(new_row.get(col, ""))

            final_rows.append(new_row)

        return final_rows

    def refresh_combined_and_final(self, source_data):
        self.populate_table(self.combined_table, source_data)
        self.final_data = self.make_final_rows(source_data)
        self.populate_table(self.final_table, self.final_data)
        self.update_dynamic_amount_labels()
        if not self.is_internal_change:
            self.save_edit_log(show_message=False)

    def hide_combined_below_level(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        try:
            level = int(self.hide_level_edit.text().strip())
            if level <= 0:
                raise ValueError
        except Exception:
            self.msg_error("錯誤", "階層請輸入大於 0 的數字，例如：4、5、6。")
            return

        self.push_edit_history()

        data = [
            row for row in self.combined_data
            if not str(row.get("項目", "")).strip()
            or len(str(row.get("項目", "")).strip()) <= level
        ]

        self.refresh_combined_and_final(data)
        self.status_label.setText(f"組合總表已隱藏第 {level + 1} 層以下資料｜顯示 {len(data)} 筆｜隱藏 {len(self.combined_data) - len(data)} 筆")

    def show_all_combined(self):
        if self.combined_data:
            self.push_edit_history()
        self.refresh_combined_and_final(self.combined_data)
        self.status_label.setText(f"組合總表已恢復全部顯示｜共 {len(self.combined_data)} 筆")

    # =========================================================
    # 顏色
    # =========================================================
    def set_cell_color(self, table, row, col_name, color_name):
        c = COLUMNS.index(col_name)
        item = table.item(row, c)
        self.set_item_foreground_silent(table, item, color_name)

    def apply_color_marks_to_combined(self):
        for (row, col_name), color in self.color_marks.items():
            self.set_cell_color(self.combined_table, row, col_name, color)

    def mark_remaining_price_amount_red(self):
        for r, row in enumerate(self.combined_data):
            if str(row.get("單價", "")).strip():
                self.color_marks[(r, "單價")] = "red"

            if str(row.get("複價", "")).strip():
                self.color_marks[(r, "複價")] = "red"

        self.apply_color_marks_to_combined()

    def mark_calc2_blue(self, changed_cells):
        item_to_row = {
            str(row.get("項目", "")).strip(): idx
            for idx, row in enumerate(self.combined_data)
            if str(row.get("項目", "")).strip()
        }

        for item, col in changed_cells:
            r = item_to_row.get(str(item).strip())
            if r is not None:
                self.color_marks[(r, col)] = "blue"

        self.apply_color_marks_to_combined()

    # =========================================================
    # 計算
    # =========================================================
    def item_map(self):
        return {
            str(row.get("項目", "")).strip(): row
            for row in self.combined_data
            if str(row.get("項目", "")).strip()
        }

    def rollup_children(self, parent, item_to_row):
        return [
            f"{parent}{code}"
            for code in CODES
            if f"{parent}{code}" in item_to_row
        ]

    def is_leaf(self, item, all_items):
        return not any(
            other != item and other.startswith(item) and len(other) > len(item)
            for other in all_items
        )

    def calculate_leaf_amounts(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        self.push_edit_history()

        all_items = [
            str(row.get("項目", "")).strip()
            for row in self.combined_data
            if str(row.get("項目", "")).strip()
        ]

        changed = 0
        skipped = 0

        for row in self.combined_data:
            item = str(row.get("項目", "")).strip()

            if not self.is_leaf(item, all_items):
                continue

            qty = self.num(row.get("數量", ""))
            price = self.num(row.get("單價", ""))

            if qty is None or price is None:
                skipped += 1
                continue

            row["複價"] = self.fmt_3(qty * price)
            changed += 1

        self.refresh_combined_and_final(self.combined_data)
        self.msg_info("完成", f"複價計算完成\n完成：{changed} 筆\n略過：{skipped} 筆")

    def calculate_rollup_amounts(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        self.push_edit_history()

        item_to_row = self.item_map()
        items = sorted(item_to_row.keys(), key=len, reverse=True)

        leaf_changed = 0
        rollup_changed = 0
        skipped = 0
        changed_cells = []

        for item in items:
            if self.rollup_children(item, item_to_row):
                continue

            row = item_to_row[item]
            qty = self.num(row.get("數量", ""))
            price = self.num(row.get("單價", ""))

            if qty is not None and price is not None:
                row["複價"] = self.fmt_3(qty * price)
                changed_cells.append((item, "複價"))
                leaf_changed += 1
            else:
                skipped += 1

        for item in items:
            children = self.rollup_children(item, item_to_row)

            if not children:
                continue

            total = 0.0
            found = False

            for child in children:
                amount = self.num(item_to_row[child].get("複價", ""))

                if amount is not None:
                    total += amount
                    found = True

            if not found:
                skipped += 1
                continue

            row = item_to_row[item]
            row["單價"] = self.fmt_3(total)
            changed_cells.append((item, "單價"))

            qty = self.num(row.get("數量", ""))
            row["複價"] = self.fmt_3(total * qty) if qty is not None else self.fmt_3(total)
            changed_cells.append((item, "複價"))
            rollup_changed += 1

        self.refresh_combined_and_final(self.combined_data)
        self.mark_calc2_blue(changed_cells)

        self.msg_info(
            "完成",
            f"重新計算完成\n"
            f"末階計算：{leaf_changed} 筆\n"
            f"階層彙總：{rollup_changed} 筆\n"
            f"略過：{skipped} 筆"
        )

    # =========================================================
    # 刪除單複價規則
    # =========================================================
    def apply_delete_rules(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        self.push_edit_history()

        unit_items = {
            str(row.get("項目", "")).strip()
            for row in self.unit_price_data
            if str(row.get("項目", "")).strip()
        }

        combined_items = [
            str(row.get("項目", "")).strip()
            for row in self.combined_data
            if str(row.get("項目", "")).strip()
        ]

        def has_child(item):
            return any(
                other != item and other.startswith(item) and len(other) > len(item)
                for other in combined_items
            )

        def has_unit_child(item):
            return any(
                other != item and other.startswith(item) and len(other) > len(item)
                for other in unit_items
            )

        def delete_price_allowed(row, item):
            if has_child(item) and has_unit_child(item):
                row["單價"] = ""

        changed = 0

        for row in self.combined_data:
            item = str(row.get("項目", "")).strip()

            if not item:
                continue

            old = (row.get("單價", ""), row.get("複價", ""))
            prefix2 = item[:2].upper()
            item_len = len(item)
            is_unit_layer = item in unit_items

            if item.startswith("011"):
                if 2 <= item_len <= 5:
                    delete_price_allowed(row, item)
                    row["複價"] = ""
                elif item_len == 6:
                    row["複價"] = ""

            elif item.startswith("012"):
                row["複價"] = ""

            elif item.startswith("013"):
                if item in {"013", "0131", "0132", "0133"}:
                    delete_price_allowed(row, item)
                elif has_unit_child(item):
                    row["單價"] = ""
                row["複價"] = ""

            elif item.startswith("014"):
                if item in {"014", "0141", "0142"}:
                    row["單價"] = ""
                    row["複價"] = ""
                elif is_unit_layer:
                    row["複價"] = ""
                elif has_unit_child(item):
                    row["單價"] = ""
                    row["複價"] = ""
                else:
                    row["複價"] = ""

            elif item == "01" or item.startswith("01"):
                delete_price_allowed(row, item)
                row["複價"] = ""

            elif self.is_02_to_0a(item):
                if item_len == 2:
                    if prefix2 in {"02", "0A"}:
                        row["單價"] = ""
                    elif prefix2 not in {"03", "04", "05", "06"}:
                        delete_price_allowed(row, item)
                    row["複價"] = ""
                else:
                    row["複價"] = ""

            if old != (row.get("單價", ""), row.get("複價", "")):
                changed += 1

        self.refresh_combined_and_final(self.combined_data)
        self.mark_remaining_price_amount_red()
        self.msg_info("完成", f"刪除單複價完成\n影響 {changed} 筆資料")

    def is_02_to_0a(self, item):
        return len(item) >= 2 and item[:2].upper() in {
            "02", "03", "04", "05", "06", "07", "08", "09", "0A"
        }

    # =========================================================
    # 整理備註
    # =========================================================
    def clean_remark_string(self, remark):
        text = "" if remark is None else str(remark)

        if text.strip() == "":
            return ""

        match = re.search(r"[#*,]", text)

        if not match:
            return text

        return text[match.start() + 1:]

    def clean_combined_remarks(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        self.push_edit_history()

        changed = 0

        for row in self.combined_data:
            old = row.get("備註", "")
            new = self.clean_remark_string(old)

            if old != new:
                row["備註"] = new
                changed += 1

        self.refresh_combined_and_final(self.combined_data)
        self.msg_info("完成", f"整理備註完成\n影響 {changed} 筆資料")

    # =========================================================
    # 輸出總表動態填入固定欄位
    # =========================================================
    def fill_final_fixed_fields(self):
        """
        將目前輸入欄位：
        - 檔別
        - 動支單號
        - 單位檢核末碼

        動態填入輸出總表所有資料列。
        """
        if not self.final_data:
            self.msg_warn("提醒", "輸出總表目前沒有資料可填入。")
            return

        self.push_edit_history()

        file_type = self.file_type_edit.text()
        payment_no = self.payment_no_edit.text()
        unit_check = self.unit_check_edit.text()

        for row in self.final_data:
            row["檔別"] = file_type
            row["動支單號"] = payment_no
            row["單位檢核末碼"] = unit_check

        self.populate_table(self.final_table, self.final_data)
        self.status_label.setText(
            f"已將檔別、動支單號、單位檢核末碼填入輸出總表｜共 {len(self.final_data)} 筆"
        )
        self.save_edit_log(show_message=False)
        self.msg_info("完成", f"已填入輸出總表固定欄位\n共 {len(self.final_data)} 筆")

    # =========================================================
    # 匯入回饋 Excel
    # =========================================================
    def normalize_import_rows(self, raw_rows):
        if not raw_rows:
            return []

        headers = [str(h).strip() for h in raw_rows[0]]
        index_map = {header: i for i, header in enumerate(headers)}
        rows = []

        for raw in raw_rows[1:]:
            if not any(str(value).strip() for value in raw):
                continue

            row = {}

            for col in COLUMNS:
                idx = index_map.get(col)
                row[col] = "" if idx is None or idx >= len(raw) else str(raw[idx]).strip()

            rows.append(row)

        return rows

    def read_html_xls_rows(self, filename):
        content = None

        for enc in ("utf-8-sig", "utf-8", "cp950", "big5"):
            try:
                content = Path(filename).read_text(encoding=enc)
                break
            except Exception:
                pass

        if content is None:
            raise ValueError("無法讀取 XLS 檔案。若這是舊版二進位 .xls，請先另存成 .xlsx 或用本程式匯出的 .xls。")

        parser = SimpleHTMLTableParser()
        parser.feed(content)

        return self.normalize_import_rows(parser.rows)

    def read_openpyxl_rows(self, filename):
        wb = load_workbook(filename, data_only=True, read_only=True)
        ws = wb.active
        raw_rows = []

        for row in ws.iter_rows(values_only=True):
            raw_rows.append(["" if cell is None else str(cell) for cell in row])

        return self.normalize_import_rows(raw_rows)

    def import_feedback_excel(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "匯入回饋 Excel", "",
            "Excel (*.xls *.xlsx *.xlsm);;All Files (*.*)"
        )

        if not filename:
            return

        try:
            lower = filename.lower()

            if lower.endswith(".xls"):
                rows = self.read_html_xls_rows(filename)
            elif lower.endswith(".xlsx") or lower.endswith(".xlsm"):
                rows = self.read_openpyxl_rows(filename)
            else:
                self.msg_error("錯誤", "只支援 .xls、.xlsx、.xlsm")
                return

            if not rows:
                self.msg_warn("提醒", "匯入檔案沒有可用資料，請確認第一列是欄位名稱。")
                return

            clean_rows = []

            for row in rows:
                new_row = {col: row.get(col, "") for col in COLUMNS}

                for col in ("數量", "單價", "複價"):
                    new_row[col] = self.final_number_text(new_row.get(col, ""))

                clean_rows.append(new_row)

            self.push_edit_history()

            self.final_data = clean_rows
            self.combined_data = [dict(row) for row in clean_rows]

            self.populate_table(self.combined_table, self.combined_data)
            self.populate_table(self.final_table, self.final_data)

            self.status_label.setText(f"已匯入回饋 Excel：{len(clean_rows)} 筆")
            self.save_edit_log(show_message=False)
            self.msg_info("完成", f"已匯入回饋 Excel：{len(clean_rows)} 筆")

        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    # =========================================================
    # 匯出 Excel
    # =========================================================
    def get_final_export_rows(self):
        if not self.final_data and self.combined_data:
            self.final_data = self.make_final_rows(self.combined_data)
            self.populate_table(self.final_table, self.final_data)

        return self.final_data

    def export_final_xls(self):
        rows = self.get_final_export_rows()

        if not rows:
            self.msg_warn("提醒", "輸出總表沒有資料可匯出。")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "匯出 XLS", "", "Excel 97-2003 (*.xls)"
        )

        if not filename:
            return

        if not filename.lower().endswith(".xls"):
            filename += ".xls"

        try:
            def html_escape(value):
                value = "" if value is None else str(value)
                return (
                    value.replace("&", "&amp;")
                         .replace("<", "&lt;")
                         .replace(">", "&gt;")
                         .replace('"', "&quot;")
                )

            output = []
            output.append('<html lang="zh-Hant-TW">')
            output.append('<head>')
            output.append('<meta charset="utf-8">')
            output.append('<meta http-equiv="Content-Type" content="text/html; charset=utf-8">')
            output.append('<meta name="ProgId" content="Excel.Sheet">')
            output.append('<style>')
            output.append('td, th { mso-number-format:"\\@"; white-space:nowrap; }')
            output.append('</style>')
            output.append('</head>')
            output.append('<body>')
            output.append('<table border="1">')
            output.append('<tr>')

            for col in COLUMNS:
                output.append(f'<th style="mso-number-format:\\@">{html_escape(col)}</th>')

            output.append('</tr>')

            for row in rows:
                output.append('<tr>')
                for col in COLUMNS:
                    output.append(f'<td style="mso-number-format:\\@">{html_escape(row.get(col, ""))}</td>')
                output.append('</tr>')

            output.append('</table></body></html>')

            Path(filename).write_text("\n".join(output), encoding="utf-8-sig")
            self.msg_info("完成", f"XLS 匯出完成：\n{filename}")

        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    def export_final_xlsm(self):
        rows = self.get_final_export_rows()

        if not rows:
            self.msg_warn("提醒", "輸出總表沒有資料可匯出。")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "匯出 XLSM", "", "Excel Macro-Enabled Workbook (*.xlsm)"
        )

        if not filename:
            return

        if not filename.lower().endswith(".xlsm"):
            filename += ".xlsm"

        try:
            wb = Workbook()
            wb.properties.creator = "預算書分析系統"
            wb.properties.title = "輸出總表"
            wb.properties.subject = "UTF-8 繁體中文輸出"

            ws = wb.active
            ws.title = "輸出總表"

            for c, col in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=1, column=c)
                cell.value = str(col)
                cell.data_type = "s"
                cell.number_format = "@"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for r, row in enumerate(rows, start=2):
                for c, col in enumerate(COLUMNS, start=1):
                    cell = ws.cell(row=r, column=c)
                    cell.value = "" if row.get(col, "") is None else str(row.get(col, ""))
                    cell.data_type = "s"
                    cell.number_format = "@"

            for c, col in enumerate(COLUMNS, start=1):
                max_len = len(str(col))
                for row in rows[:1000]:
                    max_len = max(max_len, len(str(row.get(col, ""))))
                ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(max(max_len + 2, 8), 40)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            wb.save(filename)
            self.msg_info("完成", f"XLSM 匯出完成：\n{filename}")

        except Exception:
            self.msg_error("錯誤", traceback.format_exc())


def main():
    app = QApplication(sys.argv)
    win = BudgetAnalyzerQT()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
