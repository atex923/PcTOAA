# PsTOAA_V4_5_6.py
# KOHAKU預算書轉換容器 V4.5.6
#
# 支援 PySide6 / PyQt6。
#
# 主要功能：
# - 讀取預算總表、預算詳細表、預算單價分析表。
# - 產生組合總表與輸出總表，支援刪除單複價、計算工項複價、重新計算、整理備註。
# - 支援手動編輯同步、編輯紀錄儲存/載入、回復一次、匯入回饋 Excel、匯出 XLSX。
# - 右上金額區顯示工程費、稅費、工程管理費、公共藝術費等動態金額。
# - 表格工具包含目前分頁搜尋、欄寬調整、項目名稱/備註字數檢查。
#
# 顏色標示：
# - 紅色字：刪除單複價後保留的單價/複價。
# - 紫色字：手動修改或同步修改。
# - 藍色字：重新計算產生的數字。
# - 米橘底：項目名稱或備註超過 32 字。
# - 粉綠底：第四分頁組合總表最後一層工項。
#
# 版本紀錄摘要：
# - V3.6.1 / V4.0.C1：加入內部更新鎖、編輯紀錄、手動儲存/載入、回復與右上動態金額區。
# - V4.1.x：精簡解析與表格共用邏輯，整理分頁與功能按鈕名稱。
# - V4.2.x：整合單價分析略過規則、搜尋、字數檢查、項目 0 總工程經費列、
#            右上金額區、工程管理費、匯出檔名、功能區重排、往後回復編輯與 XLSX/XLSM 輸出調整。
# - V4.3.x：整理第四分頁複價精度、0A1 到 0AA 手動單價保留、右上金額雙欄、
#            0B/0B1 計算規則、欄位清除、右鍵雙擊編輯、上半部縮放與收合開關。
# - V4.4：由 Codex Transfer 匯入，固定上半部隱藏/顯示開關文字，避免版面跳動。
# - V4.4.1：第四分頁新增整理備註2，將英文字開頭且含數字的英數混合字串標為藍色。
# - V4.4.2：整理備註2 增加數字開頭且含英文字母的英數混合字串藍色標示。
# - V4.4.3：整理備註2 執行時從備註左側刪除連續出現的英數混合特殊字串。
# - V4.4.4：合併整理備註與整理備註2，按鈕名稱改為備註整理。
# - V4.4.5：第四分頁新增刪除行按鈕，刪除游標所在的組合總表列。
# - V4.4.6：刪除行移到左側組合總表區，並讓功能執行後保留隱藏階層狀態。
# - V4.5.0：支援拖曳檔案到視窗開啟，調整勾選文字，中文匯出檔名加入版號與 TAISHOSANKE。
# - V4.5.1：修正備註整理藍字重繪、刪除行不跳完成視窗，項目 0 加總所有第二階層工項。
# - V4.5.6：建立第三碼進版規則，並讓 .pyw 與 .py 保持完整同步碼。

import sys
import re
import html
import traceback
import subprocess
import importlib
import json
import copy
from datetime import datetime
from pathlib import Path
from html.parser import HTMLParser
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation


# =========================================================
# .pyw / pythonw 無控制台模式相容處理
# =========================================================
# 在 Windows 以 .pyw 執行時，sys.stdout / sys.stderr 可能不存在。
# 若套件自動安裝或例外處理使用 print/traceback，可先導向到空裝置，避免程式因無控制台而中斷。
if sys.stdout is None:
    sys.stdout = open(Path.cwd() / "NUL", "w", encoding="utf-8", errors="ignore")
if sys.stderr is None:
    sys.stderr = open(Path.cwd() / "NUL", "w", encoding="utf-8", errors="ignore")


# =========================================================
# 缺少套件自動安裝
# =========================================================
def ensure_package(import_name, pip_name=None):
    """
    Import a dependency when running from source.

    In a Nuitka-built executable, dependencies must be bundled at build time.
    Falling back to pip install from a frozen app can call the executable
    itself as python -m pip and produce confusing errors, so frozen builds
    fail fast with a clear message instead.
    """
    pip_name = pip_name or import_name

    try:
        return importlib.import_module(import_name)
    except ImportError as import_error:
        if getattr(sys, "frozen", False):
            raise ImportError(
                f"Missing bundled dependency in executable: {import_name}. "
                f"Rebuild with Nuitka include option for {pip_name}."
            ) from import_error

        print(f"[missing package] {import_name}; installing {pip_name}")

        try:
            subprocess.check_call([
                sys.executable,
                "-m",
                "pip",
                "install",
                pip_name
            ])
        except Exception as install_error:
            print(f"[install failed] {pip_name}")
            raise install_error

        return importlib.import_module(import_name)

# Excel / 資料套件採延遲載入：
# 啟動視窗時先不載入 pandas / openpyxl / xlrd，可明顯縮短啟動時間並降低初始記憶體。
pd = None

def get_pandas():
    global pd
    if pd is None:
        pd = ensure_package("pandas", "pandas")
    return pd

def is_na_value(value):
    if value is None:
        return True
    try:
        return value != value
    except Exception:
        return False


# GUI 套件：優先使用 PySide6，若安裝失敗再嘗試 PyQt6
try:
    ensure_package("PySide6", "PySide6")

    from PySide6.QtCore import Qt, QEvent
    from PySide6.QtGui import QColor, QBrush, QTextDocument
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
        QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit,
        QPushButton, QCheckBox, QTabWidget, QTableWidget, QTableWidgetItem,
        QHeaderView, QInputDialog, QAbstractItemView, QSizePolicy,
        QStyledItemDelegate, QStyle, QStyleOptionViewItem
    )

except Exception:
    ensure_package("PyQt6", "PyQt6")

    from PyQt6.QtCore import Qt, QEvent
    from PyQt6.QtGui import QColor, QBrush, QTextDocument
    from PyQt6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
        QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit,
        QPushButton, QCheckBox, QTabWidget, QTableWidget, QTableWidgetItem,
        QHeaderView, QInputDialog, QAbstractItemView, QSizePolicy,
        QStyledItemDelegate, QStyle, QStyleOptionViewItem
    )


CODES = "123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
APP_VERSION = "V4.5.6"
APP_TITLE = f"KOHAKU預算書轉換容器{APP_VERSION}"
DEFAULT_LOG_FILENAME = "BudgetAnalyzer_V4_5_6.budget_log.json"
RIGHT_TOP_0A_CODES = ("0A1", "0A2", "0A3", "0A4", "0A5", "0A6", "0A7", "0A8", "0A9", "0AA")

COLUMNS = (
    "檔別", "動支單號", "項目", "項目名稱", "來源編號", "來源名稱",
    "序列", "材料編號", "單位", "數量", "單價", "複價",
    "比例", "計算式", "規格", "備註", "單位檢核末碼"
)

AMOUNT_COLUMNS = {"數量", "單價", "複價"}
LEFT_COLUMNS = {"項目", "項目名稱", "備註"}
EMPTY_SHRINK_COLUMNS = {
    "來源編號", "來源名稱", "序列", "材料編號",
    "比例", "計算式", "規格"
}
TEXT_LENGTH_CHECK_COLUMNS = ("項目名稱", "備註")
TEXT_LENGTH_LIMIT = 32
TEXT_LENGTH_HIGHLIGHT = "#FCE4D6"
LEAF_ROW_BACKGROUND = "#E2F0D9"
REMARK2_PATTERN = re.compile(
    r"(?<![A-Za-z0-9])"
    r"((?:[A-Za-z](?=[A-Za-z0-9]*\d)|\d(?=[A-Za-z0-9]*[A-Za-z]))[A-Za-z0-9]+)"
    r"(?![A-Za-z0-9])"
)
REMARK2_LEADING_TOKEN_PATTERN = re.compile(
    r"^\s*"
    r"(?:(?:[A-Za-z](?=[A-Za-z0-9]*\d)|\d(?=[A-Za-z0-9]*[A-Za-z]))[A-Za-z0-9]+)"
    r"(?![A-Za-z0-9])"
)
REMARK2_LEADING_SEPARATOR_PATTERN = re.compile(r"^[ \t]*[,，、;；:#＃*＊/\\_－-]+[ \t]*")


class RemarkHighlightDelegate(QStyledItemDelegate):
    def __init__(self, owner, parent=None):
        super().__init__(parent)
        self.owner = owner

    def highlighted_html(self, text):
        parts = []
        last = 0

        for match in REMARK2_PATTERN.finditer(text):
            parts.append(html.escape(text[last:match.start()]))
            parts.append(f'<span style="color:#0070C0;">{html.escape(match.group(1))}</span>')
            last = match.end()

        parts.append(html.escape(text[last:]))
        return "".join(parts).replace("\n", "<br>")

    def paint(self, painter, option, index):
        if not getattr(self.owner, "remark2_highlight_enabled", False):
            super().paint(painter, option, index)
            return

        value = index.data(Qt.ItemDataRole.DisplayRole)
        text = "" if value is None else str(value)
        if not REMARK2_PATTERN.search(text):
            super().paint(painter, option, index)
            return

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        opt.text = ""
        widget = opt.widget
        style = widget.style() if widget is not None else QApplication.style()
        style.drawControl(QStyle.ControlElement.CE_ItemViewItem, opt, painter, widget)

        doc = QTextDocument()
        doc.setDefaultFont(opt.font)
        doc.setHtml(self.highlighted_html(text))
        text_rect = style.subElementRect(QStyle.SubElement.SE_ItemViewItemText, opt, widget)

        painter.save()
        painter.translate(text_rect.topLeft())
        doc.setTextWidth(text_rect.width())
        doc.drawContents(painter)
        painter.restore()


class SimpleHTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self.current_row = None
        self.current_cell = None
        self.in_cell = False

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "tr":
            self.current_row = []
        elif tag in ("td", "th"):
            self.current_cell = []
            self.in_cell = True

    def handle_data(self, data):
        if self.in_cell and self.current_cell is not None:
            self.current_cell.append(data)

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("td", "th"):
            value = "".join(self.current_cell or [])
            value = html.unescape(value).strip()
            if self.current_row is not None:
                self.current_row.append(value)
            self.current_cell = None
            self.in_cell = False
        elif tag == "tr":
            if self.current_row is not None:
                self.rows.append(self.current_row)
            self.current_row = None


class BudgetAnalyzerQT(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle(APP_TITLE)
        screen = QApplication.primaryScreen()
        if screen is not None:
            available = screen.availableGeometry()
            self.resize(min(1700, int(available.width() * 0.96)), min(980, int(available.height() * 0.92)))
        else:
            self.resize(1500, 900)
        self.setMinimumSize(980, 680)
        self.setAcceptDrops(True)

        self.data = []
        self.detail_data = []
        self.unit_price_data = []
        self.combined_data = []
        self.final_data = []
        self.original_amounts = {}
        self.recalculated_amounts = {}

        self.color_marks = {}  # (row, col_name) -> QColor，組合總表計算/刪除顏色
        self.edited_cell_marks = {}  # (table_name, row, col_name) -> QColor，手動修改紫色

        self.is_loading_table = False
        self.is_internal_change = False
        self.edit_history = []
        self.redo_history = []
        self.max_history = 10
        self.last_exceeded_text_pos = None  # (table_index, row, col)
        self.remark2_highlight_enabled = False
        self.combined_hidden_level = None
        self.last_clear_click = None  # (row, col, timestamp_ms, count)
        self.force_single_cell_editing = False
        self.single_edit_cell = None  # (table_name, row, col)

        self.build_ui()

    # =========================================================
    # 共用工具
    # =========================================================
    def table_configs(self):
        return (
            ("summary", self.data, self.summary_table),
            ("detail", self.detail_data, self.detail_table),
            ("unit_price", self.unit_price_data, self.unit_price_table),
            ("combined", self.combined_data, self.combined_table),
            ("final", self.final_data, self.final_table),
        )

    def all_tables(self):
        return tuple(table for _, _, table in self.table_configs())

    def msg_info(self, title, text):
        QMessageBox.information(self, title, text)

    def msg_warn(self, title, text):
        QMessageBox.warning(self, title, text)

    def msg_error(self, title, text):
        QMessageBox.critical(self, title, text)

    def fmt(self, value):
        try:
            if is_na_value(value) or value == "":
                return ""
            return f"{float(value):,.2f}"
        except Exception:
            return "" if value is None else str(value)

    def fmt_2(self, value):
        """
        第四分頁「複價」金額格式：
        1. 先將計算結果整理到小數後第四位。
        2. 再依小數後第三位做四捨五入，顯示/儲存到小數後第二位。
        使用 Decimal ROUND_HALF_UP，避免 Python round 的銀行家捨入造成 0.005 類型誤差。
        """
        try:
            if is_na_value(value) or value == "":
                return ""
            text = str(value).replace(",", "").strip()
            number = Decimal(text)
            fourth = number.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
            second = fourth.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            return f"{second:,.2f}"
        except (InvalidOperation, ValueError, TypeError):
            return "" if value is None else str(value)

    def fmt_3(self, value):
        try:
            if is_na_value(value) or value == "":
                return ""
            return f"{float(value):,.3f}"
        except Exception:
            return "" if value is None else str(value)

    def num(self, value):
        try:
            text = "" if value is None else str(value).replace(",", "").strip()
            return None if text == "" else float(text)
        except Exception:
            return None

    def text(self, value):
        return "" if is_na_value(value) else str(value).strip()

    def compact(self, value):
        return "" if is_na_value(value) else "".join(str(value).strip().split())

    def is_blank(self, value):
        return self.fmt(value) == ""

    def engine_for(self, filename):
        if filename.lower().endswith(".xls"):
            ensure_package("xlrd", "xlrd")
            return "xlrd"

        ensure_package("openpyxl", "openpyxl")
        return "openpyxl"

    def record(self, item="", name="", unit="", qty="", unit_price="", amount="", remark="", price3=False, qty3=False):
        return {
            "檔別": self.file_type_edit.text(),
            "動支單號": self.payment_no_edit.text(),
            "項目": item,
            "項目名稱": name,
            "來源編號": "",
            "來源名稱": "",
            "序列": "",
            "材料編號": "",
            "單位": unit,
            "數量": self.fmt_3(qty) if qty3 else self.fmt(qty),
            "單價": self.fmt_3(unit_price),
            "複價": self.fmt_3(amount) if price3 else self.fmt(amount),
            "比例": "",
            "計算式": "",
            "規格": "",
            "備註": remark,
            "單位檢核末碼": self.unit_check_edit.text(),
        }

    # =========================================================
    # UI
    # =========================================================
    def build_ui(self):
        central = QWidget()
        central.setAcceptDrops(True)
        self.setCentralWidget(central)

        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(8, 6, 8, 6)
        main_layout.setSpacing(5)

        # 上方區域：左側操作列、中間功能區、右側動態金額資訊，三區可隨視窗縮放。
        top = QWidget()
        self.top_panel = top
        top_layout = QHBoxLayout(top)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(8)
        main_layout.addWidget(top)

        left_control_height = max(18, int(self.fontMetrics().height() * 1.2))
        right_control_height = max(18, int(24 * 0.8))

        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(3)
        left_panel.setMinimumWidth(360)
        left_panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        top_layout.addWidget(left_panel, stretch=4)

        function_panel = QWidget()
        function_panel.setMinimumWidth(240)
        function_panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        function_panel.setStyleSheet(
            "background-color: #DDEBF7; "
            "border: 1px solid #9EADCC; "
            "border-radius: 6px;"
        )
        function_layout = QVBoxLayout(function_panel)
        function_layout.setContentsMargins(10, 7, 10, 7)
        function_layout.setSpacing(5)
        top_layout.addWidget(function_panel, stretch=2)

        def compact_button(text, slot, min_width=88):
            button = QPushButton(text)
            button.clicked.connect(slot)
            button.setMinimumWidth(min_width)
            button.setFixedHeight(left_control_height)
            button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            button.setStyleSheet(
                "QPushButton {"
                "background-color: #F8FBFF; "
                "border: 1px solid #6F86A6; "
                "border-top-color: #FFFFFF; "
                "border-left-color: #FFFFFF; "
                "border-radius: 4px; "
                "padding: 2px 8px;"
                "}"
                "QPushButton:hover { background-color: #FFFFFF; }"
                "QPushButton:pressed {"
                "background-color: #C9DAF8; "
                "border-top-color: #6F86A6; "
                "border-left-color: #6F86A6; "
                "border-bottom-color: #FFFFFF; "
                "border-right-color: #FFFFFF;"
                "}"
            )
            return button

        def function_button(text, slot, min_width=88):
            return compact_button(text, slot, min_width)

        def small_label(text):
            label = QLabel(text)
            label.setStyleSheet("font-weight: bold;")
            return label

        self.file_edit = QLineEdit()
        self.project_name_edit = QLineEdit()
        self.project_no_edit = QLineEdit()
        self.file_type_edit = QLineEdit()
        self.payment_no_edit = QLineEdit()
        self.unit_check_edit = QLineEdit()
        self.hide_level_edit = QLineEdit()
        self.search_edit = QLineEdit()

        self.file_edit.setMinimumWidth(120)
        self.project_name_edit.setMinimumWidth(140)
        self.project_no_edit.setMinimumWidth(75)
        self.file_type_edit.setMinimumWidth(45)
        self.payment_no_edit.setMinimumWidth(75)
        self.unit_check_edit.setMinimumWidth(55)
        self.hide_level_edit.setMinimumWidth(42)
        self.hide_level_edit.setStyleSheet("background-color: white;")
        self.search_edit.setMinimumWidth(80)
        self.search_edit.setPlaceholderText("目前分頁搜尋")
        self.search_edit.setStyleSheet("background-color: white;")
        self.search_edit.returnPressed.connect(self.search_current_table)

        for edit in (
            self.file_edit, self.project_name_edit, self.project_no_edit,
            self.file_type_edit, self.payment_no_edit, self.unit_check_edit,
            self.hide_level_edit, self.search_edit,
        ):
            edit.setFixedHeight(left_control_height)
            edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            edit.setStyleSheet("background-color: white; border: 1px solid black;")

        # -----------------------------------------------------
        # 第1列：檔案選擇與主要操作
        # -----------------------------------------------------
        file_row = QHBoxLayout()
        file_row.setSpacing(4)
        file_row.addWidget(small_label("Excel檔案"))
        file_row.addWidget(self.file_edit, stretch=1)
        file_row.addWidget(compact_button("瀏覽", self.browse_file, 64))
        file_row.addWidget(compact_button("分析", self.analyze, 64))
        file_row.addWidget(compact_button("自動欄寬", self.auto_resize_all, 86))
        left_layout.addLayout(file_row)

        # -----------------------------------------------------
        # 第2列：工程資訊
        # -----------------------------------------------------
        project_row = QHBoxLayout()
        project_row.setSpacing(4)
        project_row.addWidget(small_label("工程名稱"))
        project_row.addWidget(self.project_name_edit, stretch=1)
        project_row.addWidget(small_label("工程編號"))
        project_row.addWidget(self.project_no_edit)
        left_layout.addLayout(project_row)

        # -----------------------------------------------------
        # 第3列：固定欄位與編輯設定
        # -----------------------------------------------------
        fixed_row = QHBoxLayout()
        fixed_row.setSpacing(4)
        fixed_row.addWidget(small_label("檔別"))
        fixed_row.addWidget(self.file_type_edit)
        fixed_row.addWidget(small_label("動支單號"))
        fixed_row.addWidget(self.payment_no_edit)
        fixed_row.addWidget(small_label("單位檢核末碼"))
        fixed_row.addWidget(self.unit_check_edit)

        self.edit_check = QCheckBox("編輯資料")
        self.edit_check.toggled.connect(self.on_edit_check_toggled)
        self.manual_width_check = QCheckBox("調整欄寬")
        self.clear_column_check = QCheckBox("欄位清除")
        self.clear_column_check.toggled.connect(self.on_clear_column_toggled)
        fixed_row.addStretch(1)
        left_layout.addLayout(fixed_row)

        # -----------------------------------------------------
        # 第4列：組合總表分析與顯示控制
        # -----------------------------------------------------
        analysis_row = QHBoxLayout()
        analysis_row.setSpacing(4)
        analysis_row.addWidget(small_label("組合總表"))
        analysis_row.addWidget(compact_button("刪除行", self.delete_current_combined_row, 72))
        analysis_row.addWidget(compact_button("刪除單複價", self.apply_delete_rules, 92))
        analysis_row.addWidget(compact_button("計算工項複價", self.calculate_leaf_amounts, 104))
        analysis_row.addWidget(compact_button("重新計算", self.calculate_rollup_amounts, 86))
        analysis_row.addStretch(1)
        left_layout.addLayout(analysis_row)

        # -----------------------------------------------------
        # 第5列：輸出總表與檔案輸出入 / 編輯紀錄
        # -----------------------------------------------------
        io_row = QHBoxLayout()
        io_row.setSpacing(4)
        io_row.addWidget(small_label("輸出總表"))
        io_row.addWidget(compact_button("填入欄位", self.fill_final_fixed_fields, 82))
        io_row.addStretch(1)
        left_layout.addLayout(io_row)

        def add_function_row(*widgets):
            row = QHBoxLayout()
            row.setSpacing(4)
            for widget in widgets:
                row.addWidget(widget)
            row.addStretch(1)
            function_layout.addLayout(row)

        add_function_row(
            small_label("搜尋"),
            self.search_edit,
            function_button("下一筆", self.search_current_table, 64),
            function_button("清除", self.clear_search, 58),
        )
        add_function_row(
            small_label("隱藏階層"),
            self.hide_level_edit,
            function_button("隱藏", self.hide_combined_below_level, 58),
            function_button("恢復", self.show_all_combined, 58),
        )
        add_function_row(self.edit_check, self.clear_column_check, self.manual_width_check)
        add_function_row(
            function_button("備註整理", self.organize_combined_remarks, 88),
            function_button("檢查字數", self.check_text_lengths, 82),
            function_button("尋找超長", self.find_next_exceeded_text, 82),
        )
        add_function_row(
            function_button("儲存編輯", self.save_edit_log_manual, 82),
            function_button("載入編輯", self.load_edit_log_manual, 82),
        )
        add_function_row(
            function_button("匯出 XLSX", self.export_final_xlsx, 88),
            function_button("匯入回饋", self.import_feedback_excel, 82),
        )
        add_function_row(
            function_button("往前回復編輯", self.undo_last_edit, 104),
            function_button("往後回復編輯", self.redo_last_edit, 104),
        )

        # -----------------------------------------------------
        # 右側：動態金額資訊區
        # -----------------------------------------------------
        self.amount_panel = QWidget()
        self.amount_panel.setMinimumWidth(520)
        self.amount_panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.amount_panel.setStyleSheet(
            "background-color: #FFF2CC; "
            "border: 1px solid #D6B656; "
            "border-radius: 8px;"
        )
        amount_panel_layout = QGridLayout(self.amount_panel)
        amount_panel_layout.setContentsMargins(12, 6, 12, 6)
        amount_panel_layout.setHorizontalSpacing(5)
        amount_panel_layout.setVerticalSpacing(2)
        amount_panel_layout.setColumnStretch(0, 3)
        amount_panel_layout.setColumnStretch(1, 2)
        amount_panel_layout.setColumnStretch(2, 1)
        amount_panel_layout.setColumnStretch(3, 1)
        amount_panel_layout.setColumnStretch(4, 2)

        # 右上資訊區欄位較多，字體與列高縮小到約 80% 避免資料擁擠。
        name_style = "color: red; font-weight: bold; font-size: 10px; background-color: transparent;"
        value_style = "color: red; font-weight: bold; font-size: 10px; background-color: transparent;"
        calc_name_style = "color: red; font-weight: bold; font-size: 10px; background-color: transparent;"

        self.amount_01_name_label = QLabel("01發包工程費")
        self.amount_011_name_label = QLabel("011包工程")
        self.amount_0b_name_label = QLabel("0B營業稅")
        self.amount_profit_name_label = QLabel("012承包商利潤及工程保險費")
        self.amount_air_name_label = QLabel("0A1空氣污染防制費")
        self.amount_management_name_label = QLabel("06工程管理費")
        self.amount_qc_name_label = QLabel("0A2二級品管抽驗費")
        self.amount_art_name_label = QLabel("0A3公共藝術費")
        self.amount_total_name_label = QLabel("總經費")

        self.amount_0b_calc_name_label = QLabel("01×5%")
        self.amount_profit_calc_name_label = QLabel("011×10%")
        self.amount_air_calc_name_label = QLabel("01×0.5%")
        self.amount_management_calc_name_label = QLabel("級距試算")
        self.amount_qc_calc_name_label = QLabel("01×0.1%")
        self.amount_art_calc_name_label = QLabel("01X1.0%")
        self.amount_01_calc_name_label = QLabel("重新計算")
        self.amount_011_calc_name_label = QLabel("重新計算")
        self.amount_total_calc_name_label = QLabel("重新計算")

        self.amount_01_value_label = QLabel("0.000")
        self.amount_011_value_label = QLabel("0.000")
        self.amount_0b_value_label = QLabel("0.000")
        self.amount_0b_calc_value_label = QLabel("0.000")
        self.amount_profit_value_label = QLabel("0.000")
        self.amount_profit_calc_value_label = QLabel("0.000")
        self.amount_air_value_label = QLabel("0.000")
        self.amount_air_calc_value_label = QLabel("0.000")
        self.amount_management_value_label = QLabel("0.000")
        self.amount_management_calc_value_label = QLabel("0")
        self.amount_qc_value_label = QLabel("0.000")
        self.amount_qc_calc_value_label = QLabel("0.000")
        self.amount_art_summary_value_label = QLabel("0.000")
        self.amount_art_combined_value_label = QLabel("0.000")
        self.amount_total_value_label = QLabel("0.000")
        self.amount_01_calc_value_label = QLabel("0.000")
        self.amount_011_calc_value_label = QLabel("0.000")
        self.amount_total_calc_value_label = QLabel("0.000")

        self.amount_0a_rows = {
            "0A1": {
                "name": self.amount_air_name_label,
                "actual": self.amount_air_value_label,
                "source": QLineEdit("01"),
                "ratio": QLineEdit("0.005"),
                "result": self.amount_air_calc_value_label,
            },
            "0A2": {
                "name": self.amount_qc_name_label,
                "actual": self.amount_qc_value_label,
                "source": QLineEdit("01"),
                "ratio": QLineEdit("0.001"),
                "result": self.amount_qc_calc_value_label,
            },
            "0A3": {
                "name": self.amount_art_name_label,
                "actual": self.amount_art_summary_value_label,
                "source": QLineEdit("01"),
                "ratio": QLineEdit("0.01"),
                "result": self.amount_art_combined_value_label,
            },
        }

        for code in RIGHT_TOP_0A_CODES[3:]:
            self.amount_0a_rows[code] = {
                "name": QLabel(code),
                "actual": QLabel("0.000"),
                "source": QLineEdit(""),
                "ratio": QLineEdit("1"),
                "result": QLabel("0.000"),
            }

        name_labels = [
            self.amount_01_name_label,
            self.amount_011_name_label,
            self.amount_0b_name_label,
            self.amount_profit_name_label,
            self.amount_air_name_label,
            self.amount_management_name_label,
            self.amount_qc_name_label,
            self.amount_art_name_label,
            self.amount_total_name_label,
        ]
        calc_name_labels = [
            self.amount_01_calc_name_label,
            self.amount_011_calc_name_label,
            self.amount_0b_calc_name_label,
            self.amount_profit_calc_name_label,
            self.amount_air_calc_name_label,
            self.amount_management_calc_name_label,
            self.amount_qc_calc_name_label,
            self.amount_art_calc_name_label,
            self.amount_total_calc_name_label,
        ]
        value_labels = [
            self.amount_01_value_label,
            self.amount_011_value_label,
            self.amount_0b_value_label,
            self.amount_0b_calc_value_label,
            self.amount_profit_value_label,
            self.amount_profit_calc_value_label,
            self.amount_air_value_label,
            self.amount_air_calc_value_label,
            self.amount_management_value_label,
            self.amount_management_calc_value_label,
            self.amount_qc_value_label,
            self.amount_qc_calc_value_label,
            self.amount_art_summary_value_label,
            self.amount_art_combined_value_label,
            self.amount_total_value_label,
            self.amount_01_calc_value_label,
            self.amount_011_calc_value_label,
            self.amount_total_calc_value_label,
        ]

        for code in RIGHT_TOP_0A_CODES[3:]:
            name_labels.append(self.amount_0a_rows[code]["name"])
            value_labels.append(self.amount_0a_rows[code]["actual"])
            value_labels.append(self.amount_0a_rows[code]["result"])

        for label in name_labels:
            label.setStyleSheet(name_style)
            label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            label.setFixedHeight(right_control_height)
            label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)

        for label in calc_name_labels:
            label.setStyleSheet(calc_name_style)
            label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            label.setFixedHeight(right_control_height)
            label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)

        for label in value_labels:
            label.setStyleSheet(value_style)
            label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            label.setFixedHeight(right_control_height)
            label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)

        for row in self.amount_0a_rows.values():
            for edit in (row["source"], row["ratio"]):
                edit.setMinimumWidth(42)
                edit.setFixedHeight(right_control_height)
                edit.setStyleSheet("background-color: white; border: 1px solid #9E9E9E; font-size: 10px;")
                edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
                edit.editingFinished.connect(self.update_dynamic_amount_labels)

        self.amount_panel.setMinimumHeight(220)
        for label in (
            self.amount_01_value_label, self.amount_01_calc_value_label,
            self.amount_011_value_label, self.amount_011_calc_value_label,
            self.amount_total_value_label, self.amount_total_calc_value_label,
            self.amount_0b_value_label, self.amount_0b_calc_value_label,
            self.amount_profit_value_label, self.amount_profit_calc_value_label,
            self.amount_air_value_label, self.amount_air_calc_value_label,
            self.amount_management_value_label, self.amount_management_calc_value_label,
            self.amount_qc_value_label, self.amount_qc_calc_value_label,
            self.amount_art_summary_value_label, self.amount_art_combined_value_label,
        ):
            label.setMinimumWidth(92)
            label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

        for code in RIGHT_TOP_0A_CODES[3:]:
            for label in (self.amount_0a_rows[code]["actual"], self.amount_0a_rows[code]["result"]):
                label.setMinimumWidth(92)
                label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

        # 雙欄列：左邊保留分析讀入原始複價，右邊顯示重新計算後結果。
        dual_rows = [
            (0, self.amount_01_name_label, self.amount_01_value_label, self.amount_01_calc_name_label, self.amount_01_calc_value_label),
            (1, self.amount_011_name_label, self.amount_011_value_label, self.amount_011_calc_name_label, self.amount_011_calc_value_label),
            (2, self.amount_profit_name_label, self.amount_profit_value_label, self.amount_profit_calc_name_label, self.amount_profit_calc_value_label),
            (3, self.amount_0b_name_label, self.amount_0b_value_label, self.amount_0b_calc_name_label, self.amount_0b_calc_value_label),
            (4, self.amount_management_name_label, self.amount_management_value_label, self.amount_management_calc_name_label, self.amount_management_calc_value_label),
        ]

        for row_no, title_label, actual_label, calc_label, calc_value_label in dual_rows:
            amount_panel_layout.addWidget(title_label, row_no, 0)
            amount_panel_layout.addWidget(actual_label, row_no, 1)
            amount_panel_layout.addWidget(calc_label, row_no, 2, 1, 2)
            amount_panel_layout.addWidget(calc_value_label, row_no, 4)

        for row_no, code in enumerate(RIGHT_TOP_0A_CODES, start=5):
            row = self.amount_0a_rows[code]
            amount_panel_layout.addWidget(row["name"], row_no, 0)
            amount_panel_layout.addWidget(row["actual"], row_no, 1)
            amount_panel_layout.addWidget(row["source"], row_no, 2)
            amount_panel_layout.addWidget(row["ratio"], row_no, 3)
            amount_panel_layout.addWidget(row["result"], row_no, 4)

        total_row_no = 5 + len(RIGHT_TOP_0A_CODES)
        amount_panel_layout.addWidget(self.amount_total_name_label, total_row_no, 0)
        amount_panel_layout.addWidget(self.amount_total_value_label, total_row_no, 1)
        amount_panel_layout.addWidget(self.amount_total_calc_name_label, total_row_no, 2, 1, 2)
        amount_panel_layout.addWidget(self.amount_total_calc_value_label, total_row_no, 4)

        self.amount_total_name_label.setStyleSheet(
            name_style + "border-top: 1px solid #D6B656; padding-top: 3px;"
        )
        self.amount_total_value_label.setStyleSheet(
            value_style + "border-top: 1px solid #D6B656; padding-top: 3px;"
        )
        self.amount_total_calc_name_label.setStyleSheet(
            calc_name_style + "border-top: 1px solid #D6B656; padding-top: 3px;"
        )
        self.amount_total_calc_value_label.setStyleSheet(
            value_style + "border-top: 1px solid #D6B656; padding-top: 3px;"
        )

        top_layout.addWidget(self.amount_panel, stretch=5)

        # -----------------------------------------------------
        # 上半部收合 / 展開按鈕：放在上半部與下半部分頁表格中間。
        # -----------------------------------------------------
        toggle_row = QHBoxLayout()
        toggle_row.setContentsMargins(0, 2, 0, 2)
        toggle_row.addStretch(1)

        self.top_toggle_button = QPushButton("上半部隱藏/顯示開關")
        self.top_toggle_button.setFixedHeight(24)
        self.top_toggle_button.setMinimumWidth(150)
        self.top_toggle_button.setToolTip("暫時隱藏或重新顯示上半部功能區")
        self.top_toggle_button.clicked.connect(self.toggle_top_panel)
        self.top_toggle_button.setStyleSheet(
            "QPushButton {"
            "background-color: #E2F0D9; "
            "border: 1px solid #70AD47; "
            "border-radius: 6px; "
            "font-weight: bold;"
            "}"
            "QPushButton:hover { background-color: #F2F8EE; }"
            "QPushButton:pressed { background-color: #C6E0B4; }"
        )

        toggle_row.addWidget(self.top_toggle_button)
        toggle_row.addStretch(1)
        main_layout.addLayout(toggle_row)

        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs, stretch=1)

        self.summary_table = self.create_table()
        self.detail_table = self.create_table()
        self.unit_price_table = self.create_table()
        self.combined_table = self.create_table()
        self.final_table = self.create_table()
        self.combined_table.setItemDelegateForColumn(
            COLUMNS.index("備註"),
            RemarkHighlightDelegate(self, self.combined_table)
        )

        self.tabs.addTab(self.summary_table, "預算總表")
        self.tabs.addTab(self.detail_table, "預算詳細表")
        self.tabs.addTab(self.unit_price_table, "預算單價分析表")
        self.tabs.addTab(self.combined_table, "組合總表")
        self.tabs.addTab(self.final_table, "輸出總表")

        self.status_label = QLabel("尚未分析")
        main_layout.addWidget(self.status_label)

        self.slogan_label = QLabel("Imagination is more important than knowledge.")
        self.slogan_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.slogan_label.setStyleSheet("color: #555555; font-size: 12px;")
        main_layout.addWidget(self.slogan_label)

    def toggle_top_panel(self):
        """暫時隱藏或重新顯示上半部功能區。"""
        hidden = self.top_panel.isVisible()
        self.top_panel.setVisible(not hidden)

        if hidden:
            self.top_toggle_button.setText("上半部隱藏/顯示開關")
            self.top_toggle_button.setToolTip("上半部目前已顯示，按下可隱藏")
        else:
            self.top_toggle_button.setText("上半部隱藏/顯示開關")
            self.top_toggle_button.setToolTip("上半部目前已隱藏，按下可顯示")


    def create_table(self):
        table = QTableWidget()
        table.setColumnCount(len(COLUMNS))
        table.setHorizontalHeaderLabels(COLUMNS)
        table.setShowGrid(True)
        table.setGridStyle(Qt.PenStyle.SolidLine)
        table.setAlternatingRowColors(False)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.verticalHeader().setVisible(True)
        table.itemChanged.connect(self.on_item_changed)
        table.installEventFilter(self)
        table.viewport().installEventFilter(self)
        return table

    def on_clear_column_toggled(self, checked):
        if checked and self.edit_check.isChecked():
            self.edit_check.setChecked(False)
        self.update_editable_state()

    def on_edit_check_toggled(self, checked):
        if checked and hasattr(self, "clear_column_check") and self.clear_column_check.isChecked():
            self.clear_column_check.setChecked(False)
        self.update_editable_state()

    def eventFilter(self, obj, event):
        # 注意：viewport/table 會收到很多事件，必須先判斷事件類型，再讀取對應屬性。
        if hasattr(self, "combined_table") and obj in (self.combined_table, self.combined_table.viewport()):
            event_type = event.type()

            if event_type == QEvent.Type.KeyPress and obj is self.combined_table:
                if (
                    event.key() == Qt.Key.Key_Delete
                    and getattr(self, "clear_column_check", None) is not None
                    and self.clear_column_check.isChecked()
                ):
                    self.clear_selected_combined_cells()
                    return True
                return super().eventFilter(obj, event)

            if obj is not self.combined_table.viewport():
                return super().eventFilter(obj, event)

            mouse_events = (QEvent.Type.MouseButtonPress, QEvent.Type.MouseButtonDblClick)
            if event_type not in mouse_events:
                return super().eventFilter(obj, event)

            button = event.button()
            index = self.combined_table.indexAt(event.pos())

            if not index.isValid():
                return False

            clear_mode = (
                getattr(self, "clear_column_check", None) is not None
                and self.clear_column_check.isChecked()
            )

            if (
                event_type == QEvent.Type.MouseButtonPress
                and button == Qt.MouseButton.LeftButton
                and clear_mode
                and event.modifiers() & Qt.KeyboardModifier.ShiftModifier
            ):
                self.clear_combined_cell(index.row(), index.column())
                return True

            if event_type == QEvent.Type.MouseButtonDblClick and button == Qt.MouseButton.RightButton:
                self.start_combined_cell_right_double_edit(index.row(), index.column())
                return True

        return super().eventFilter(obj, event)

    def clear_selected_combined_cells(self):
        indexes = self.combined_table.selectedIndexes()

        if not indexes:
            current_row = self.combined_table.currentRow()
            current_col = self.combined_table.currentColumn()
            if current_row >= 0 and current_col >= 0:
                indexes = [self.combined_table.model().index(current_row, current_col)]

        cells = sorted({(idx.row(), idx.column()) for idx in indexes if idx.isValid()})

        if not cells:
            self.status_label.setText("欄位清除模式｜沒有選取欄位")
            return

        self.clear_combined_cells(cells)

    def clear_combined_cell(self, table_row, col):
        self.clear_combined_cells([(table_row, col)])

    def clear_combined_cells(self, cells):
        changed = []

        for table_row, col in cells:
            data_row = self.data_index_for_table_row(self.combined_table, self.combined_data, table_row)
            if data_row is None or not (0 <= data_row < len(self.combined_data)):
                continue

            if col < 0 or col >= len(COLUMNS):
                continue

            col_name = COLUMNS[col]
            old_value = "" if self.combined_data[data_row].get(col_name, "") is None else str(self.combined_data[data_row].get(col_name, ""))
            if old_value == "":
                continue

            changed.append((table_row, data_row, col_name))

        if not changed:
            self.status_label.setText("欄位清除模式｜選取欄位已是空白")
            return

        self.push_edit_history()

        for table_row, data_row, col_name in changed:
            self.combined_data[data_row][col_name] = ""
            self.set_table_cell_value_silent(self.combined_table, table_row, col_name, "")
            self.mark_manual_cell("combined", data_row, col_name)
            item_code = str(self.combined_data[data_row].get("項目", "")).strip()
            self.sync_edit_to_other_pages("combined", data_row, item_code, col_name, "")

        self.final_data = self.make_final_rows(self.combined_data)
        self.populate_table(self.final_table, self.final_data)
        self.update_dynamic_amount_labels()
        self.save_edit_log(show_message=False)
        self.status_label.setText(f"欄位清除完成｜{len(changed)} 格")

    def shift_color_marks_after_row_delete(self, deleted_row):
        shifted = {}

        for (row, col_name), color in self.color_marks.items():
            if row == deleted_row:
                continue

            shifted[(row - 1 if row > deleted_row else row, col_name)] = color

        self.color_marks = shifted

    def shift_edited_marks_after_row_delete(self, table_name, deleted_row):
        shifted = {}

        for (mark_table, row, col_name), color in self.edited_cell_marks.items():
            if mark_table == table_name:
                if row == deleted_row:
                    continue

                row = row - 1 if row > deleted_row else row

            shifted[(mark_table, row, col_name)] = color

        self.edited_cell_marks = shifted

    def delete_current_combined_row(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        table_row = self.combined_table.currentRow()

        if table_row < 0:
            self.msg_warn("提醒", "請先在組合總表點選要刪除的行。")
            return

        data_row = self.data_index_for_table_row(self.combined_table, self.combined_data, table_row)

        if data_row is None or not (0 <= data_row < len(self.combined_data)):
            self.msg_warn("提醒", "找不到游標所在行對應的資料。")
            return

        self.push_edit_history()
        deleted = self.combined_data.pop(data_row)
        item_code = str(deleted.get("項目", "")).strip()

        self.shift_color_marks_after_row_delete(data_row)
        self.shift_edited_marks_after_row_delete("combined", data_row)
        self.shift_edited_marks_after_row_delete("final", data_row)
        self.refresh_combined_and_final(self.combined_data)

        next_row = min(table_row, self.combined_table.rowCount() - 1)
        if next_row >= 0:
            self.combined_table.setCurrentCell(next_row, COLUMNS.index("項目"))

        self.status_label.setText(f"已刪除組合總表第 {data_row + 1} 列｜項目 {item_code or '空白'}")

    def start_combined_cell_right_double_edit(self, table_row, col):
        item = self.combined_table.item(table_row, col)
        if item is None:
            item = QTableWidgetItem("")
            self.combined_table.setItem(table_row, col, item)

        flags = item.flags()
        item.setFlags(flags | Qt.ItemFlag.ItemIsEditable)
        self.force_single_cell_editing = True
        self.single_edit_cell = ("combined", table_row, col)
        self.combined_table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        self.combined_table.setCurrentCell(table_row, col)
        self.combined_table.editItem(item)
        self.status_label.setText(f"右鍵雙擊編輯｜第 {table_row + 1} 列｜{COLUMNS[col]}")

    def table_data_ref(self, table):
        for name, data, target_table in self.table_configs():
            if table is target_table:
                return name, data
        return "", []

    def commit_active_editor(self):
        focus_widget = QApplication.focusWidget()

        if focus_widget is not None:
            focus_widget.clearFocus()
            QApplication.processEvents()

    def data_index_for_table_row(self, table, data, table_row):
        if table.rowCount() == len(data):
            return table_row if 0 <= table_row < len(data) else None

        item_col = COLUMNS.index("項目")
        item = table.item(table_row, item_col)
        item_code = "" if item is None else item.text().strip().upper()

        if not item_code:
            return table_row if 0 <= table_row < len(data) else None

        for idx, row in enumerate(data):
            if str(row.get("項目", "")).strip().upper() == item_code:
                return idx

        return table_row if 0 <= table_row < len(data) else None

    def sync_table_widget_to_data(self, table, data):
        synced_rows = set()

        for r in range(table.rowCount()):
            data_row = self.data_index_for_table_row(table, data, r)

            if data_row is None or data_row in synced_rows:
                continue

            for c, col in enumerate(COLUMNS):
                item = table.item(r, c)

                if item is not None:
                    data[data_row][col] = item.text()

            synced_rows.add(data_row)

    def populate_table(self, table, data):
        self.is_loading_table = True
        table.setRowCount(0)
        table.setRowCount(len(data))

        for r, row in enumerate(data):
            for c, col in enumerate(COLUMNS):
                value = "" if row.get(col, "") is None else str(row.get(col, ""))
                if table in (self.unit_price_table, self.combined_table) and col == "數量":
                    number = self.num(value)
                    if number is not None:
                        value = self.fmt_3(number)
                item = QTableWidgetItem(value)

                if col in LEFT_COLUMNS:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                elif col in AMOUNT_COLUMNS:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                self.set_item_editable_flag(item)

                table.setItem(r, c, item)

        self.is_loading_table = False

        self.apply_manual_edit_marks_to_table(table)

        if table is self.combined_table:
            self.apply_leaf_row_background_to_combined()
            self.apply_color_marks_to_combined()
            self.update_dynamic_amount_labels()
            if self.remark2_highlight_enabled:
                table.viewport().update()

        if not self.manual_width_check.isChecked():
            self.auto_resize_table(table)

        self.update_editable_state()

    def auto_resize_table(self, table):
        for c, col in enumerate(COLUMNS):
            if table is self.combined_table and col in EMPTY_SHRINK_COLUMNS:
                is_empty = True
                for r in range(table.rowCount()):
                    item = table.item(r, c)
                    if item and item.text().strip():
                        is_empty = False
                        break
                if is_empty:
                    table.setColumnWidth(c, 1)
                    continue

            table.resizeColumnToContents(c)
            width = table.columnWidth(c)
            if col in LEFT_COLUMNS:
                table.setColumnWidth(c, min(max(width + 20, 80), 700))
            else:
                table.setColumnWidth(c, min(max(width + 20, 70), 260))

    def auto_resize_all(self):
        for table in self.all_tables():
            self.auto_resize_table(table)
        self.msg_info("完成", "已自動調整欄寬")

    def search_current_table(self):
        keyword = self.search_edit.text().strip()

        if not keyword:
            self.msg_warn("提醒", "請輸入搜尋關鍵字。")
            return

        table = self.tabs.currentWidget()

        if not isinstance(table, QTableWidget) or table.rowCount() == 0:
            self.msg_warn("提醒", "目前分頁沒有可搜尋的資料。")
            return

        keyword_lower = keyword.lower()
        row_count = table.rowCount()
        col_count = table.columnCount()
        total_cells = row_count * col_count
        current_row = table.currentRow()
        current_col = table.currentColumn()

        if current_row < 0 or current_col < 0:
            start_index = 0
        else:
            start_index = (current_row * col_count + current_col + 1) % total_cells

        for offset in range(total_cells):
            index = (start_index + offset) % total_cells
            row = index // col_count
            col = index % col_count
            item = table.item(row, col)

            if item and keyword_lower in item.text().lower():
                table.setCurrentCell(row, col)
                table.scrollToItem(item, QAbstractItemView.ScrollHint.PositionAtCenter)
                col_name = COLUMNS[col] if col < len(COLUMNS) else str(col + 1)
                self.status_label.setText(f"搜尋「{keyword}」｜第 {row + 1} 列｜{col_name}")
                return

        self.msg_warn("搜尋結果", f"目前分頁找不到：{keyword}")

    def clear_search(self):
        self.search_edit.clear()
        table = self.tabs.currentWidget()

        if isinstance(table, QTableWidget):
            table.clearSelection()

        self.status_label.setText("已清除搜尋")

    def set_item_background_silent(self, table, item, color_name):
        if item is None:
            return

        old_block_state = table.blockSignals(True)
        old_internal_state = self.is_internal_change
        self.is_internal_change = True

        try:
            item.setBackground(QBrush(QColor(color_name)))
        finally:
            self.is_internal_change = old_internal_state
            table.blockSignals(old_block_state)

    def check_text_lengths(self):
        checked = 0
        exceeded = 0

        for table in self.all_tables():
            table_checked, table_exceeded = self.apply_text_length_highlights_to_table(table)
            checked += table_checked
            exceeded += table_exceeded

        self.status_label.setText(
            f"字數檢查完成｜檢查 {checked} 格｜超過 {TEXT_LENGTH_LIMIT} 字：{exceeded} 格"
        )
        self.msg_info(
            "完成",
            f"字數檢查完成\n"
            f"檢查欄位：項目名稱、備註\n"
            f"超過 {TEXT_LENGTH_LIMIT} 字：{exceeded} 格"
        )

    def base_background_for_cell(self, table, row, col):
        if table is self.combined_table and self.is_displayed_combined_leaf_row(row):
            return LEAF_ROW_BACKGROUND
        return "white"

    def apply_text_length_highlights_to_table(self, table):
        checked = 0
        exceeded = 0

        for col_name in TEXT_LENGTH_CHECK_COLUMNS:
            col_index = COLUMNS.index(col_name)

            for row in range(table.rowCount()):
                item = table.item(row, col_index)

                if item is None:
                    continue

                checked += 1
                text = item.text().strip()

                if len(text) > TEXT_LENGTH_LIMIT:
                    self.set_item_background_silent(table, item, TEXT_LENGTH_HIGHLIGHT)
                    exceeded += 1
                else:
                    self.set_item_background_silent(table, item, self.base_background_for_cell(table, row, col_index))

        return checked, exceeded

    def find_next_exceeded_text(self):
        tables = list(self.all_tables())
        table = self.tabs.currentWidget()

        if not isinstance(table, QTableWidget):
            self.msg_warn("提醒", "目前分頁沒有可搜尋的資料。")
            return

        table_index = tables.index(table) if table in tables else 0
        col_indexes = [COLUMNS.index(col_name) for col_name in TEXT_LENGTH_CHECK_COLUMNS]

        start_row = 0
        start_col_pos = 0
        if self.last_exceeded_text_pos and self.last_exceeded_text_pos[0] == table_index:
            _, last_row, last_col = self.last_exceeded_text_pos
            try:
                last_col_pos = col_indexes.index(last_col)
            except ValueError:
                last_col_pos = -1
            start_row = last_row
            start_col_pos = last_col_pos + 1
            if start_col_pos >= len(col_indexes):
                start_col_pos = 0
                start_row += 1

        row_count = table.rowCount()
        if row_count == 0:
            self.msg_warn("提醒", "目前分頁沒有資料。")
            return

        total_checks = row_count * len(col_indexes)
        start_index = ((start_row % row_count) * len(col_indexes)) + start_col_pos

        for offset in range(total_checks):
            index = (start_index + offset) % total_checks
            row = index // len(col_indexes)
            col = col_indexes[index % len(col_indexes)]
            item = table.item(row, col)
            text = "" if item is None else item.text().strip()

            if len(text) > TEXT_LENGTH_LIMIT:
                self.last_exceeded_text_pos = (table_index, row, col)
                table.setCurrentCell(row, col)
                table.scrollToItem(item, QAbstractItemView.ScrollHint.PositionAtCenter)
                self.status_label.setText(
                    f"找到超過 {TEXT_LENGTH_LIMIT} 字｜第 {row + 1} 列｜{COLUMNS[col]}｜{len(text)} 字"
                )
                return

        self.last_exceeded_text_pos = None
        self.msg_warn("搜尋結果", f"目前分頁找不到超過 {TEXT_LENGTH_LIMIT} 字的項目名稱或備註。")

    def update_editable_state(self):
        """
        切換表格可編輯狀態。
        更新 flags 時暫停 signals，避免 Qt 將內部狀態更新視為使用者編輯。
        """
        editable = self.edit_check.isChecked()

        self.is_internal_change = True

        try:
            for table in self.all_tables():
                old_block_state = table.blockSignals(True)

                try:
                    if editable:
                        table.setEditTriggers(
                            QAbstractItemView.EditTrigger.DoubleClicked
                            | QAbstractItemView.EditTrigger.EditKeyPressed
                            | QAbstractItemView.EditTrigger.AnyKeyPressed
                        )
                    else:
                        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

                    for r in range(table.rowCount()):
                        for c in range(table.columnCount()):
                            item = table.item(r, c)

                            if item is None:
                                continue

                            flags = item.flags()

                            if editable:
                                item.setFlags(flags | Qt.ItemFlag.ItemIsEditable)
                            else:
                                item.setFlags(flags & ~Qt.ItemFlag.ItemIsEditable)

                finally:
                    table.blockSignals(old_block_state)

        finally:
            self.is_internal_change = False

    def set_item_editable_flag(self, item):
        """
        建立儲存格時依目前勾選狀態設定是否可編輯。
        """
        if self.edit_check.isChecked():
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        else:
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

    def format_dynamic_amount(self, value):
        """
        上半部金額顯示格式：#,##0.000。
        若資料空白或不是數字，顯示 0.000。
        """
        number = self.num(value)
        if number is None:
            number = 0.0
        return f"{number:,.3f}"

    def get_combined_amount_by_item(self, item_code):
        fallback = ""

        for row in self.combined_data:
            if str(row.get("項目", "")).strip().upper() == str(item_code).strip().upper():
                amount = row.get("複價", "")
                if str(amount).strip():
                    return amount
                fallback = amount

        return fallback

    def get_summary_amount_by_item(self, item_code):
        for row in self.data:
            if str(row.get("項目", "")).strip().upper() == str(item_code).strip().upper():
                return row.get("複價", "")
        return ""

    def get_combined_row_by_item(self, item_code):
        target = str(item_code).strip().upper()

        for row in self.combined_data:
            if str(row.get("項目", "")).strip().upper() == target:
                return row

        return None

    def get_combined_name_by_item(self, item_code):
        row = self.get_combined_row_by_item(item_code)
        return "" if row is None else str(row.get("項目名稱", "")).strip()

    def amount_0a_config(self, code):
        row = self.amount_0a_rows.get(code)

        if not row:
            return "", None, ""

        source = row["source"].text().strip().upper()
        ratio_text = row["ratio"].text().strip()
        ratio = self.num(ratio_text)

        return source, ratio, ratio_text

    def lower_total_amount(self):
        return sum(
            self.num(self.get_combined_amount_by_item(code)) or 0.0
            for code in ("01", "02", "03", "04", "05", "06", "0A", "0B")
        )

    def update_0a_amount_rows(self):
        for code, row in self.amount_0a_rows.items():
            name = self.get_combined_name_by_item(code)
            exists = self.get_combined_row_by_item(code) is not None
            visible = code in {"0A1", "0A2", "0A3"} or exists

            row["name"].setText(f"{code}{name}" if name else code)

            for widget in row.values():
                widget.setVisible(visible)

    def calculate_management_fee(self, base_amount):
        brackets = (
            (5_000_000, 0.03),
            (20_000_000, 0.015),
            (75_000_000, 0.01),
            (400_000_000, 0.007),
        )
        remaining = max(base_amount, 0.0)
        total = 0.0

        for limit, rate in brackets:
            amount = min(remaining, limit)
            total += amount * rate
            remaining -= amount

            if remaining <= 0:
                return total

        return total + remaining * 0.005

    def dynamic_amount_values(self):
        amounts = {
            code: self.num(self.get_combined_amount_by_item(code)) or 0.0
            for code in (
                "0", "01", "011", "0B", "012", "02", "03", "04", "05", "06", "0A",
                *RIGHT_TOP_0A_CODES
            )
        }
        amount_0a3_summary = self.num(self.get_summary_amount_by_item("0A3")) or 0.0
        base_01 = amounts["01"]
        base_011 = amounts["011"]
        management_base = amounts["011"] + sum(amounts[code] for code in ("02", "03", "04", "05"))
        management_fee = self.calculate_management_fee(management_base)
        calculated = {
            "012": base_011 * 0.10,
            "0B": base_01 * 0.05,
            "06": int(management_fee),
        }

        for code in RIGHT_TOP_0A_CODES:
            source, ratio, ratio_text = self.amount_0a_config(code)

            if ratio == 0:
                calculated[code] = 0.0
            elif source and ratio is not None:
                calculated[code] = (self.num(self.get_combined_amount_by_item(source)) or 0.0) * ratio
            elif not source and ratio == 1:
                row = self.get_combined_row_by_item(code)
                qty = self.num(row.get("數量", "")) if row is not None else None
                unit_price = self.num(row.get("單價", "")) if row is not None else None
                calculated[code] = qty * unit_price if qty is not None and unit_price is not None else 0.0
            else:
                calculated[code] = 0.0

        return {
            "amounts": amounts,
            "amount_0a3_summary": amount_0a3_summary,
            "base_01": base_01,
            "base_011": base_011,
            "management_fee": management_fee,
            "calculated": calculated,
        }

    def display_amount_snapshot(self, calculated_side=False):
        values = self.dynamic_amount_values()
        amounts = values["amounts"]
        calculated = values["calculated"]
        total_amount = self.lower_total_amount()

        if calculated_side:
            snapshot = {
                "01": amounts["01"],
                "011": amounts["011"],
                "012": calculated["012"],
                "0B": calculated["0B"],
                "06": calculated["06"],
                "0A1": calculated["0A1"],
                "0A2": calculated["0A2"],
                "0A3": calculated["0A3"],
                "total": total_amount,
            }
            for code in RIGHT_TOP_0A_CODES[3:]:
                snapshot[code] = calculated[code]
            return snapshot

        original_snapshot = {
            "01": amounts["01"],
            "011": amounts["011"],
            "012": amounts["012"],
            "0B": amounts["0B"],
            "06": amounts["06"],
            "0A1": amounts["0A1"],
            "0A2": amounts["0A2"],
            "0A3": values["amount_0a3_summary"],
            "total": total_amount,
        }
        for code in RIGHT_TOP_0A_CODES[3:]:
            original_snapshot[code] = amounts[code]
        return original_snapshot

    def get_combined_amount_by_name(self, item_name):
        """
        依組合總表「項目名稱」抓複價。
        先做完全相同比對；找不到時再用包含比對，提高不同預算書命名的容錯。
        """
        target = self.compact(item_name)

        for row in self.combined_data:
            name = self.compact(row.get("項目名稱", ""))
            if name == target:
                return row.get("複價", "")

        for row in self.combined_data:
            name = self.compact(row.get("項目名稱", ""))
            if target and (target in name or name in target):
                return row.get("複價", "")

        return ""

    def update_dynamic_amount_labels(self):
        """
        更新右上金額區。
        左欄多取組合總表複價；試算欄依各項目規則計算。
        """
        required_labels = (
            "amount_01_value_label",
            "amount_01_calc_value_label",
            "amount_011_value_label",
            "amount_011_calc_value_label",
            "amount_0b_value_label",
            "amount_0b_calc_value_label",
            "amount_profit_value_label",
            "amount_profit_calc_value_label",
            "amount_air_value_label",
            "amount_air_calc_value_label",
            "amount_management_value_label",
            "amount_management_calc_value_label",
            "amount_qc_value_label",
            "amount_qc_calc_value_label",
            "amount_art_summary_value_label",
            "amount_art_combined_value_label",
            "amount_total_value_label",
            "amount_total_calc_value_label",
        )

        if not all(hasattr(self, name) for name in required_labels):
            return

        self.update_0a_amount_rows()
        original_values = self.original_amounts or self.display_amount_snapshot(calculated_side=False)
        if self.recalculated_amounts:
            self.recalculated_amounts = self.display_amount_snapshot(calculated_side=True)
        recalculated_values = self.recalculated_amounts

        label_values = [
            (self.amount_01_value_label, original_values.get("01", 0.0)),
            (self.amount_01_calc_value_label, recalculated_values.get("01", 0.0)),
            (self.amount_011_value_label, original_values.get("011", 0.0)),
            (self.amount_011_calc_value_label, recalculated_values.get("011", 0.0)),
            (self.amount_profit_value_label, original_values.get("012", 0.0)),
            (self.amount_profit_calc_value_label, recalculated_values.get("012", 0.0)),
            (self.amount_0b_value_label, original_values.get("0B", 0.0)),
            (self.amount_0b_calc_value_label, recalculated_values.get("0B", 0.0)),
            (self.amount_management_value_label, original_values.get("06", 0.0)),
            (self.amount_management_calc_value_label, recalculated_values.get("06", 0.0)),
            (self.amount_air_value_label, original_values.get("0A1", 0.0)),
            (self.amount_air_calc_value_label, recalculated_values.get("0A1", 0.0)),
            (self.amount_qc_value_label, original_values.get("0A2", 0.0)),
            (self.amount_qc_calc_value_label, recalculated_values.get("0A2", 0.0)),
            (self.amount_art_summary_value_label, original_values.get("0A3", 0.0)),
            (self.amount_art_combined_value_label, recalculated_values.get("0A3", 0.0)),
            (self.amount_total_value_label, original_values.get("total", 0.0)),
            (self.amount_total_calc_value_label, recalculated_values.get("total", 0.0)),
        ]

        for code in RIGHT_TOP_0A_CODES[3:]:
            row = self.amount_0a_rows[code]
            label_values.append((row["actual"], original_values.get(code, 0.0)))
            label_values.append((row["result"], recalculated_values.get(code, 0.0)))

        for label, value in label_values:
            label.setText(f"{value:,.3f}")

    def table_object_by_name(self, name):
        for table_name, _, table in self.table_configs():
            if table_name == name:
                return table
        return None

    def set_table_cell_value_silent(self, table, row, col_name, value):
        """
        靜默更新表格指定儲存格，避免觸發遞迴 itemChanged。
        """
        col_index = COLUMNS.index(col_name)
        old_block_state = table.blockSignals(True)

        try:
            item = table.item(row, col_index)

            if item is None:
                item = QTableWidgetItem("")
                table.setItem(row, col_index, item)

            item.setText("" if value is None else str(value))

        finally:
            table.blockSignals(old_block_state)

    def set_item_foreground_silent(self, table, item, color_name):
        if item is None:
            return

        old_block_state = table.blockSignals(True)
        old_internal_state = self.is_internal_change
        self.is_internal_change = True

        try:
            item.setForeground(QBrush(QColor(color_name)))
        finally:
            self.is_internal_change = old_internal_state
            table.blockSignals(old_block_state)

    def mark_manual_cell(self, table_name, row, col_name):
        """
        只標記手動修改的單一儲存格紫色。
        染色屬於程式內部視覺更新，必須暫停 signals，避免再次觸發 itemChanged。
        """
        self.edited_cell_marks[(table_name, row, col_name)] = "purple"

        table = self.table_object_by_name(table_name)

        if table is None:
            return

        col_index = COLUMNS.index(col_name)
        item = table.item(row, col_index)

        self.set_item_foreground_silent(table, item, "purple")

    def apply_manual_edit_marks_to_table(self, table):
        """
        表格重新載入後，恢復手動修改儲存格的紫色標記。
        """
        name, _ = self.table_data_ref(table)

        for (table_name, row, col_name), color in list(self.edited_cell_marks.items()):
            if table_name != name:
                continue

            if row >= table.rowCount():
                continue

            col_index = COLUMNS.index(col_name)
            item = table.item(row, col_index)

            if item:
                item.setForeground(QBrush(QColor(color)))

    def sync_edit_to_other_pages(self, source_name, source_row, item_code, col_name, value):
        """
        只要有一個分頁修改資料：
        依「項目」欄位，在其他分頁找到相同項目，更新同一欄位資料。
        組合總表/輸出總表互相同步；其他分頁也會同步相同項目。
        """
        if not item_code:
            return

        for target_name, target_data, target_table in self.table_configs():
            for idx, row in enumerate(target_data):
                if target_name == source_name and idx == source_row:
                    continue

                if str(row.get("項目", "")).strip() != item_code:
                    continue

                row[col_name] = value
                self.set_table_cell_value_silent(target_table, idx, col_name, value)

                # 同步更新的格子也標紫色，代表這格是因手動修改而變更。
                self.mark_manual_cell(target_name, idx, col_name)

    def on_item_changed(self, item):
        if self.is_loading_table or self.is_internal_change:
            return

        table = self.sender()

        if not isinstance(table, QTableWidget):
            return

        name, data = self.table_data_ref(table)

        allow_single_edit = self.force_single_cell_editing and name == "combined"

        if not name or (not self.edit_check.isChecked() and not allow_single_edit):
            return

        r = item.row()
        c = item.column()
        col = COLUMNS[c]
        value = item.text()
        data_row = self.data_index_for_table_row(table, data, r)

        old_value = ""
        if data_row is not None:
            old_value = "" if data[data_row].get(col, "") is None else str(data[data_row].get(col, ""))
            if old_value == value:
                return
            self.push_edit_history()
            data[data_row][col] = value

        # 只讓被手動修改的這一格變成紫色。
        self.mark_manual_cell(name, r, col)

        # 依「項目」同步其他分頁相同項目的同欄位。
        item_code = ""
        if data_row is not None:
            item_code = str(data[data_row].get("項目", "")).strip()

        self.sync_edit_to_other_pages(name, data_row if data_row is not None else r, item_code, col, value)
        self.update_dynamic_amount_labels()
        self.save_edit_log(show_message=False)

        if allow_single_edit and self.single_edit_cell == (name, r, c):
            self.force_single_cell_editing = False
            self.single_edit_cell = None
            self.update_editable_state()


    # =========================================================
    # V4.2.7 儲存 / 載入 / 編輯歷史
    # =========================================================
    def log_path_for_current_file(self):
        """
        依目前 Excel 檔案產生專用 log 檔名。
        例：AAA.xlsx -> AAA.budget_log.json
        """
        filename = self.file_edit.text().strip()

        if filename:
            path = Path(filename)
            if path.parent.exists():
                return path.with_suffix(".budget_log.json")

        return Path.cwd() / DEFAULT_LOG_FILENAME

    def serialize_marks(self, marks):
        output = []

        for key, color in marks.items():
            output.append({"key": list(key), "color": str(color)})

        return output

    def deserialize_marks(self, rows):
        output = {}

        for row in rows or []:
            key = tuple(row.get("key", []))
            color = row.get("color", "")

            if key and color:
                output[key] = color

        return output

    def collect_state(self):
        return {
            "version": APP_VERSION,
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "file_path": self.file_edit.text(),
            "project_name": self.project_name_edit.text(),
            "project_no": self.project_no_edit.text(),
            "file_type": self.file_type_edit.text(),
            "payment_no": self.payment_no_edit.text(),
            "unit_check": self.unit_check_edit.text(),
            "hide_level": self.hide_level_edit.text(),
            "combined_hidden_level": self.combined_hidden_level,
            "amount_0a_settings": {
                code: {
                    "source": row["source"].text(),
                    "ratio": row["ratio"].text(),
                }
                for code, row in self.amount_0a_rows.items()
            },
            "data": copy.deepcopy(self.data),
            "detail_data": copy.deepcopy(self.detail_data),
            "unit_price_data": copy.deepcopy(self.unit_price_data),
            "combined_data": copy.deepcopy(self.combined_data),
            "final_data": copy.deepcopy(self.final_data),
            "original_amounts": copy.deepcopy(self.original_amounts),
            "recalculated_amounts": copy.deepcopy(self.recalculated_amounts),
            "remark2_highlight_enabled": self.remark2_highlight_enabled,
            "color_marks": self.serialize_marks(self.color_marks),
            "edited_cell_marks": self.serialize_marks(self.edited_cell_marks),
        }

    def apply_state(self, state):
        self.is_internal_change = True

        try:
            self.file_edit.setText(state.get("file_path", ""))
            self.project_name_edit.setText(state.get("project_name", ""))
            self.project_no_edit.setText(state.get("project_no", ""))
            self.file_type_edit.setText(state.get("file_type", ""))
            self.payment_no_edit.setText(state.get("payment_no", ""))
            self.unit_check_edit.setText(state.get("unit_check", ""))
            self.hide_level_edit.setText(state.get("hide_level", ""))
            self.combined_hidden_level = state.get("combined_hidden_level", None)

            for code, settings in state.get("amount_0a_settings", {}).items():
                row = self.amount_0a_rows.get(code)
                if not row:
                    continue
                row["source"].setText(settings.get("source", ""))
                row["ratio"].setText(settings.get("ratio", "1"))

            self.data = copy.deepcopy(state.get("data", []))
            self.detail_data = copy.deepcopy(state.get("detail_data", []))
            self.unit_price_data = copy.deepcopy(state.get("unit_price_data", []))
            self.combined_data = copy.deepcopy(state.get("combined_data", []))
            self.final_data = copy.deepcopy(state.get("final_data", []))
            self.original_amounts = copy.deepcopy(state.get("original_amounts", {}))
            self.recalculated_amounts = copy.deepcopy(state.get("recalculated_amounts", {}))
            self.remark2_highlight_enabled = bool(state.get("remark2_highlight_enabled", False))
            self.color_marks = self.deserialize_marks(state.get("color_marks", []))
            self.edited_cell_marks = self.deserialize_marks(state.get("edited_cell_marks", []))

            self.populate_table(self.summary_table, self.data)
            self.populate_table(self.detail_table, self.detail_data)
            self.populate_table(self.unit_price_table, self.unit_price_data)
            self.populate_table(self.combined_table, self.displayed_combined_data())
            self.populate_table(self.final_table, self.final_data)
            self.update_dynamic_amount_labels()

            self.status_label.setText(
                f"已載入編輯紀錄｜預算總表：{len(self.data)} 筆｜"
                f"預算詳細表：{len(self.detail_data)} 筆｜"
                f"單價分析：{len(self.unit_price_data)} 筆｜"
                f"組合表：{len(self.combined_data)} 筆｜"
                f"輸出總表：{len(self.final_data)} 筆"
            )

        finally:
            self.is_internal_change = False
            self.update_editable_state()

    def make_history_snapshot(self):
        state = self.collect_state()
        state["history_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return state

    def push_edit_history(self):
        self.edit_history.append(self.make_history_snapshot())
        self.redo_history = []

        if len(self.edit_history) > self.max_history:
            self.edit_history = self.edit_history[-self.max_history:]

    def save_edit_log(self, show_message=False):
        try:
            log_path = self.log_path_for_current_file()
            state = self.collect_state()
            payload = {
                "app": "BudgetAnalyzerQT",
                "version": APP_VERSION,
                "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "state": state,
                "history": self.edit_history[-self.max_history:],
                "redo_history": self.redo_history[-self.max_history:],
            }

            log_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )

            self.status_label.setText(f"已儲存編輯紀錄：{log_path}")

            if show_message:
                self.msg_info("完成", f"已儲存編輯紀錄：\n{log_path}")

            return True

        except Exception:
            if show_message:
                self.msg_error("錯誤", traceback.format_exc())
            return False

    def save_edit_log_manual(self):
        self.save_edit_log(show_message=True)

    def load_edit_log_from_path(self, log_path):
        payload = json.loads(Path(log_path).read_text(encoding="utf-8"))
        state = payload.get("state", payload)
        self.edit_history = payload.get("history", [])[-self.max_history:]
        self.redo_history = payload.get("redo_history", [])[-self.max_history:]
        self.apply_state(state)

    def load_edit_log_manual(self):
        default_path = str(self.log_path_for_current_file())
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "載入編輯紀錄",
            default_path,
            "Budget Analyzer Log (*.budget_log.json *.json);;All Files (*.*)"
        )

        if not filename:
            return

        try:
            self.load_edit_log_from_path(filename)
            self.msg_info("完成", f"已載入編輯紀錄：\n{filename}")
        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    def ask_load_existing_log(self):
        log_path = self.log_path_for_current_file()

        if not log_path.exists():
            return

        reply = QMessageBox.question(
            self,
            "發現編輯紀錄",
            f"偵測到此檔案有編輯紀錄：\n{log_path}\n\n是否載入上次編輯內容？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.load_edit_log_from_path(log_path)
            except Exception:
                self.msg_error("錯誤", traceback.format_exc())

    def undo_last_edit(self):
        if not self.edit_history:
            self.msg_warn("提醒", "目前沒有可回復的編輯歷史。")
            return

        try:
            state = self.edit_history.pop()
            self.redo_history.append(self.make_history_snapshot())

            if len(self.redo_history) > self.max_history:
                self.redo_history = self.redo_history[-self.max_history:]

            self.apply_state(state)
            self.save_edit_log(show_message=False)
            self.msg_info("完成", f"已回復上一次編輯。\n剩餘可回復次數：{len(self.edit_history)}")
        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    def redo_last_edit(self):
        if not self.redo_history:
            self.msg_warn("提醒", "目前沒有可往後回復的編輯歷史。")
            return

        try:
            state = self.redo_history.pop()
            self.edit_history.append(self.make_history_snapshot())

            if len(self.edit_history) > self.max_history:
                self.edit_history = self.edit_history[-self.max_history:]

            self.apply_state(state)
            self.save_edit_log(show_message=False)
            self.msg_info("完成", f"已往後回復一次。\n剩餘可往後回復次數：{len(self.redo_history)}")
        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    def closeEvent(self, event):
        self.save_edit_log(show_message=False)
        event.accept()

    # =========================================================
    # 讀取分析
    # =========================================================
    def set_selected_file(self, path):
        if not path:
            return

        self.file_edit.setText(path)
        self.file_edit.setToolTip(path)
        self.file_edit.setCursorPosition(len(path))
        self.ask_load_existing_log()

    def dragEnterEvent(self, event):
        mime = event.mimeData()

        if mime.hasUrls():
            for url in mime.urls():
                path = url.toLocalFile()
                if path.lower().endswith((".xlsx", ".xls")):
                    event.acceptProposedAction()
                    return

        event.ignore()

    def dropEvent(self, event):
        mime = event.mimeData()

        if not mime.hasUrls():
            event.ignore()
            return

        for url in mime.urls():
            path = url.toLocalFile()

            if path.lower().endswith((".xlsx", ".xls")):
                self.set_selected_file(path)
                self.status_label.setText(f"已拖曳載入檔案：{path}")
                event.acceptProposedAction()
                return

        self.msg_warn("提醒", "請拖曳 .xlsx 或 .xls 檔案。")
        event.ignore()

    def browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "選擇預算書 Excel 檔案", "",
            "Excel (*.xlsx *.xls);;All Files (*.*)"
        )
        if path:
            self.set_selected_file(path)

    def clear_loaded_data(self):
        self.data.clear()
        self.detail_data.clear()
        self.unit_price_data.clear()
        self.combined_data.clear()
        self.final_data.clear()
        self.original_amounts = {}
        self.recalculated_amounts = {}
        self.remark2_highlight_enabled = False
        self.combined_hidden_level = None
        self.color_marks.clear()
        self.edited_cell_marks.clear()
        self.edit_history = []
        self.redo_history = []

        for table in self.all_tables():
            table.setRowCount(0)

        self.update_dynamic_amount_labels()
        self.status_label.setText("已清除既有資料")

    def analyze(self):
        try:
            filename = self.file_edit.text().strip()

            if not filename:
                self.msg_warn("提醒", "請先選擇 Excel 檔案")
                return

            if any((self.data, self.detail_data, self.unit_price_data, self.combined_data, self.final_data)):
                reply = QMessageBox.question(
                    self,
                    "確認重新分析",
                    "下半部目前已有資料。\n是否全部清除並重新載入資料？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )

                if reply != QMessageBox.StandardButton.Yes:
                    return

                self.clear_loaded_data()

            engine = self.engine_for(filename)

            self.parse_summary(filename, engine)
            self.parse_detail(filename, engine)
            warnings = self.parse_unit_price(filename, engine)
            self.build_combined_data()
            self.original_amounts = self.display_amount_snapshot(calculated_side=False)
            self.recalculated_amounts = {}

            self.populate_table(self.summary_table, self.data)
            self.populate_table(self.detail_table, self.detail_data)
            self.populate_table(self.unit_price_table, self.unit_price_data)
            self.refresh_combined_and_final(self.combined_data)

            self.status_label.setText(
                f"分析完成｜預算總表：{len(self.data)} 筆｜"
                f"預算詳細表：{len(self.detail_data)} 筆｜"
                f"預算單價分析表：{len(self.unit_price_data)} 筆｜"
                f"組合表：{len(self.combined_data)} 筆｜"
                f"輸出總表：{len(self.final_data)} 筆"
            )

            self.edit_history = []
            self.redo_history = []
            self.save_edit_log(show_message=False)

            if warnings:
                self.msg_warn("單價分析材料超過 35 筆", "\n".join(warnings[:20]) + ("\n..." if len(warnings) > 20 else ""))
            else:
                self.msg_info("完成", "分析完成")

        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    def parse_budget_rows(self, df, target, has_unit_price_fields):
        target.clear()
        current = None

        for r in range(8, len(df)):
            item = self.text(df.iloc[r, 0])
            name = self.text(df.iloc[r, 1])
            remark = self.text(df.iloc[r, 6])

            if item and not item[0].isdigit():
                break

            if item:
                if current:
                    target.append(self.record(**current))

                current = {
                    "item": item,
                    "name": self.compact(name),
                    "unit": self.text(df.iloc[r, 2]) if has_unit_price_fields else "",
                    "qty": df.iloc[r, 3] if has_unit_price_fields else "",
                    "unit_price": df.iloc[r, 4] if has_unit_price_fields else "",
                    "amount": df.iloc[r, 5],
                    "remark": self.compact(remark),
                }

            elif current:
                current["name"] += self.compact(name)
                current["remark"] += self.compact(remark)

        if current:
            target.append(self.record(**current))

    def parse_summary(self, filename, engine):
        df = get_pandas().read_excel(filename, sheet_name="預算總表", header=None, engine=engine)
        self.project_name_edit.setText(self.text(df.iloc[5, 1]))
        self.project_no_edit.setText(self.text(df.iloc[6, 5]))
        self.parse_budget_rows(df, self.data, False)

    def parse_detail(self, filename, engine):
        df = get_pandas().read_excel(filename, sheet_name="預算詳細表", header=None, engine=engine)
        self.parse_budget_rows(df, self.detail_data, True)

    def parse_unit_price(self, filename, engine):
        df = get_pandas().read_excel(filename, sheet_name="預算單價分析表", header=None, engine=engine)
        self.unit_price_data.clear()
        warnings = []

        r = 7

        while r < len(df):
            parent = self.text(df.iloc[r, 0])

            if not parent:
                r += 1
                continue

            if not parent.startswith("0"):
                break

            parent_unit = self.text(df.iloc[r, 2])

            material_count = 0
            current = None
            m = r + 2

            total_qty = None
            scan = m

            while scan < len(df):
                scan_name = self.text(df.iloc[scan, 1])
                if scan_name == "合計":
                    total_qty = self.num(df.iloc[scan, 3])
                    break
                scan += 1

            def calc_material_qty(raw_qty, raw_unit_price):
                if total_qty is None or total_qty == 0:
                    return raw_qty

                if parent_unit == "式":
                    base = self.num(raw_unit_price)
                else:
                    base = self.num(raw_qty)

                if base is None:
                    return raw_qty

                return base / total_qty

            def flush():
                nonlocal material_count, current

                if not current:
                    return

                material_count += 1

                if material_count > 35:
                    warnings.append(f"工項 {parent} 的單價分析材料超過 35 筆，第 {material_count} 筆起未匯入。")
                else:
                    current["item"] = f"{parent}{CODES[material_count - 1]}"
                    self.unit_price_data.append(self.record(**current, price3=True, qty3=True))

                current = None

            def is_labor_machine_marker(value):
                marker = self.compact(value).replace("：", ":")
                return "人工:機具:" in marker

            def is_work_item_marker(value):
                marker = self.compact(value).replace("：", ":")
                return "工作項目:" in marker

            skip_labor_rows = False

            while m < len(df):
                a_value = self.text(df.iloc[m, 0])
                name = self.text(df.iloc[m, 1])
                unit = self.text(df.iloc[m, 2])
                raw_qty = df.iloc[m, 3]
                unit_price = df.iloc[m, 4]
                amount = df.iloc[m, 5]
                remark = self.text(df.iloc[m, 6])

                if name == "合計":
                    flush()
                    break

                if skip_labor_rows:
                    if a_value:
                        skip_labor_rows = False
                    else:
                        m += 1
                        continue

                if is_labor_machine_marker(name):
                    flush()
                    skip_labor_rows = True
                    m += 1
                    continue

                if is_work_item_marker(name):
                    m += 1
                    continue

                # 若讀到表頭列，略過：
                # C欄 = 單位 或 D欄 = 數量
                if unit == "單位" or self.text(raw_qty) == "數量":
                    m += 1
                    continue

                blank_row = (
                    not a_value and not name and not unit
                    and self.is_blank(raw_qty) and self.is_blank(unit_price)
                    and self.is_blank(amount) and not remark
                )

                if blank_row:
                    m += 1
                    continue

                qty = calc_material_qty(raw_qty, unit_price)

                if current and not unit:
                    current["name"] += self.compact(name)
                    if remark:
                        current["remark"] += self.compact(remark)
                    if self.is_blank(current["qty"]) and not self.is_blank(qty):
                        current["qty"] = qty
                    if self.is_blank(current["unit_price"]) and not self.is_blank(unit_price):
                        current["unit_price"] = unit_price
                    if self.is_blank(current["amount"]) and not self.is_blank(amount):
                        current["amount"] = amount
                else:
                    flush()
                    current = {
                        "item": "",
                        "name": self.compact(name),
                        "unit": unit,
                        "qty": qty,
                        "unit_price": unit_price,
                        "amount": amount,
                        "remark": self.compact(remark),
                    }

                m += 1

            flush()
            r = m + 1

        return warnings

    # =========================================================
    # 組合總表 / 輸出總表
    # =========================================================
    def build_combined_data(self):
        self.combined_data.clear()
        self.combined_data.append(self.total_project_row())

        detail_items = [
            str(row.get("項目", "")).strip()
            for row in self.detail_data
            if str(row.get("項目", "")).strip()
        ]

        child_map = {item: [] for item in detail_items}
        detail_item_set = set(detail_items)
        unmatched = []

        for row in self.unit_price_data:
            item = str(row.get("項目", "")).strip()
            parent = next(
                (item[:length] for length in range(len(item) - 1, 0, -1) if item[:length] in detail_item_set),
                ""
            )

            row = dict(row)
            row["來源名稱"] = ""

            if parent:
                child_map.setdefault(parent, []).append(row)
            else:
                unmatched.append(row)

        for row in self.detail_data:
            parent = str(row.get("項目", "")).strip()
            new_row = dict(row)
            new_row["來源名稱"] = ""
            self.combined_data.append(new_row)
            self.combined_data.extend(child_map.get(parent, []))

        for row in unmatched:
            row["來源名稱"] = ""
            self.combined_data.append(row)

        self.apply_detail_unit_prices_to_0a_items()

    def apply_detail_unit_prices_to_0a_items(self):
        detail_row_by_item = {
            str(row.get("項目", "")).strip().upper(): row
            for row in self.detail_data
            if str(row.get("項目", "")).strip().upper() in RIGHT_TOP_0A_CODES
        }

        for row in self.combined_data:
            item = str(row.get("項目", "")).strip().upper()

            if item in detail_row_by_item:
                detail_row = detail_row_by_item[item]

                if str(detail_row.get("數量", "")).strip():
                    row["數量"] = detail_row.get("數量", "")

                if str(detail_row.get("單價", "")).strip():
                    row["單價"] = detail_row.get("單價", "")

    def total_project_row(self):
        return self.record(
            item="0",
            name="總工程經費",
            unit="式",
            qty=1,
            unit_price="",
            amount="",
            remark="",
        )

    def final_number_text(self, value):
        text = "" if value is None else str(value).strip()
        if text == "":
            return ""
        return text.replace(",", "")

    def make_final_rows(self, source_data):
        final_rows = []

        for row in source_data:
            new_row = dict(row)
            new_row["檔別"] = self.file_type_edit.text()
            new_row["動支單號"] = self.payment_no_edit.text()
            new_row["單位檢核末碼"] = self.unit_check_edit.text()

            for col in ("數量", "單價", "複價"):
                new_row[col] = self.final_number_text(new_row.get(col, ""))

            final_rows.append(new_row)

        return final_rows

    def displayed_combined_data(self):
        if self.combined_hidden_level is None:
            return self.combined_data

        level = self.combined_hidden_level
        return [
            row for row in self.combined_data
            if not str(row.get("項目", "")).strip()
            or len(str(row.get("項目", "")).strip()) <= level
        ]

    def refresh_combined_and_final(self, source_data):
        if source_data is not self.combined_data:
            self.combined_hidden_level = None

        self.populate_table(self.combined_table, self.displayed_combined_data())
        self.final_data = self.make_final_rows(self.combined_data)
        self.populate_table(self.final_table, self.final_data)
        self.update_dynamic_amount_labels()
        if not self.is_internal_change:
            self.save_edit_log(show_message=False)

    def hide_combined_below_level(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        try:
            level = int(self.hide_level_edit.text().strip())
            if level <= 0:
                raise ValueError
        except Exception:
            self.msg_error("錯誤", "階層請輸入大於 0 的數字，例如：4、5、6。")
            return

        self.push_edit_history()

        self.combined_hidden_level = level
        data = self.displayed_combined_data()
        self.refresh_combined_and_final(self.combined_data)
        self.status_label.setText(f"組合總表已隱藏第 {level + 1} 層以下資料｜顯示 {len(data)} 筆｜隱藏 {len(self.combined_data) - len(data)} 筆")

    def show_all_combined(self):
        if self.combined_data:
            self.push_edit_history()
        self.combined_hidden_level = None
        self.refresh_combined_and_final(self.combined_data)
        self.status_label.setText(f"組合總表已恢復全部顯示｜共 {len(self.combined_data)} 筆")

    # =========================================================
    # 顏色
    # =========================================================
    def displayed_combined_items(self):
        items = []
        item_col = COLUMNS.index("項目")
        for row in range(self.combined_table.rowCount()):
            item = self.combined_table.item(row, item_col)
            code = "" if item is None else item.text().strip()
            if code:
                items.append(code)
        return items

    def is_displayed_combined_leaf_row(self, row):
        item_col = COLUMNS.index("項目")
        item = self.combined_table.item(row, item_col)
        code = "" if item is None else item.text().strip()
        if not code:
            return False
        items = self.displayed_combined_items()
        return self.is_leaf(code, items)

    def apply_leaf_row_background_to_combined(self):
        items = self.displayed_combined_items()
        for row in range(self.combined_table.rowCount()):
            item_widget = self.combined_table.item(row, COLUMNS.index("項目"))
            code = "" if item_widget is None else item_widget.text().strip()
            if not code or not self.is_leaf(code, items):
                continue
            for col in range(self.combined_table.columnCount()):
                cell = self.combined_table.item(row, col)
                if cell is not None:
                    self.set_item_background_silent(self.combined_table, cell, LEAF_ROW_BACKGROUND)

    def set_cell_color(self, table, row, col_name, color_name):
        c = COLUMNS.index(col_name)
        item = table.item(row, c)
        self.set_item_foreground_silent(table, item, color_name)

    def apply_color_marks_to_combined(self):
        for (row, col_name), color in self.color_marks.items():
            self.set_cell_color(self.combined_table, row, col_name, color)

    def mark_remaining_price_amount_red(self):
        for r, row in enumerate(self.combined_data):
            if str(row.get("單價", "")).strip():
                self.color_marks[(r, "單價")] = "red"

            if str(row.get("複價", "")).strip():
                self.color_marks[(r, "複價")] = "red"

        self.apply_color_marks_to_combined()

    def mark_calc2_blue(self, changed_cells):
        item_to_row = {
            str(row.get("項目", "")).strip(): idx
            for idx, row in enumerate(self.combined_data)
            if str(row.get("項目", "")).strip()
        }

        for item, col in changed_cells:
            r = item_to_row.get(str(item).strip())
            if r is not None:
                self.color_marks[(r, col)] = "blue"

        self.apply_color_marks_to_combined()

    # =========================================================
    # 計算
    # =========================================================
    def item_map(self):
        return {
            str(row.get("項目", "")).strip(): row
            for row in self.combined_data
            if str(row.get("項目", "")).strip()
        }

    def rollup_children(self, parent, item_to_row):
        return [
            f"{parent}{code}"
            for code in CODES
            if f"{parent}{code}" in item_to_row
        ]

    def sync_dynamic_calculations_to_combined(self, item_to_row):
        target_codes = ("012", "06", *RIGHT_TOP_0A_CODES)
        changed_cells = []
        changed = 0

        for item_code in target_codes:
            row = item_to_row.get(item_code)

            if row is None:
                continue

            unit_price = self.num(row.get("單價", ""))

            if unit_price is None:
                continue

            qty = self.num(row.get("數量", ""))
            if qty is None:
                continue

            row["數量"] = self.fmt_3(qty)
            row["複價"] = self.fmt_2(unit_price * qty)

            changed_cells.append((item_code, "複價"))
            changed += 1

        return changed, changed_cells

    def rollup_parent_amounts(self, item_to_row, changed_cells):
        rollup_changed = 0
        skipped = 0
        items = sorted(item_to_row.keys(), key=len, reverse=True)

        for item in items:
            children = self.rollup_children(item, item_to_row)

            if not children:
                continue

            if item.upper() in RIGHT_TOP_0A_CODES:
                continue

            total = 0.0
            found = False

            for child in children:
                amount = self.num(item_to_row[child].get("複價", ""))

                if amount is not None:
                    total += amount
                    found = True

            if not found:
                skipped += 1
                continue

            row = item_to_row[item]
            row["單價"] = self.fmt_3(total)
            changed_cells.append((item, "單價"))

            qty = self.num(row.get("數量", ""))
            row["複價"] = self.fmt_2(total * qty) if qty is not None else self.fmt_2(total)
            changed_cells.append((item, "複價"))
            rollup_changed += 1

        return rollup_changed, skipped

    def calculate_total_project_from_second_level(self, item_to_row, changed_cells):
        row = item_to_row.get("0")

        if row is None:
            return 0

        total = 0.0
        found = False

        for item, child_row in item_to_row.items():
            code = str(item).strip().upper()

            if len(code) != 2 or not code.startswith("0"):
                continue

            amount = self.num(child_row.get("複價", ""))

            if amount is None:
                continue

            total += amount
            found = True

        if not found:
            return 0

        row["單價"] = self.fmt_3(total)
        row["複價"] = self.fmt_2(total)
        changed_cells.extend([("0", "單價"), ("0", "複價")])
        return 1

    def is_leaf(self, item, all_items):
        return not any(
            other != item and other.startswith(item) and len(other) > len(item)
            for other in all_items
        )

    def calculate_leaf_amounts(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        self.commit_active_editor()
        self.sync_table_widget_to_data(self.combined_table, self.combined_data)
        self.push_edit_history()

        all_items = [
            str(row.get("項目", "")).strip()
            for row in self.combined_data
            if str(row.get("項目", "")).strip()
        ]

        changed = 0
        skipped = 0

        for row in self.combined_data:
            item = str(row.get("項目", "")).strip()

            if item.upper() in {"0B", "0B1"}:
                continue

            if not self.is_leaf(item, all_items):
                continue

            qty = self.num(row.get("數量", ""))
            price = self.num(row.get("單價", ""))

            if qty is None or price is None:
                skipped += 1
                continue

            row["複價"] = self.fmt_2(qty * price)
            changed += 1

        self.refresh_combined_and_final(self.combined_data)
        self.msg_info("完成", f"複價計算完成\n完成：{changed} 筆\n略過：{skipped} 筆")

    def calculate_0b1_and_0b(self, item_to_row, changed_cells):
        changed = 0
        row_01 = item_to_row.get("01")
        row_0b1 = item_to_row.get("0B1")
        row_0b = item_to_row.get("0B")

        base_01 = self.num(row_01.get("複價", "")) if row_01 is not None else None

        if row_0b1 is not None and base_01 is not None:
            tax_unit_price = base_01 * 0.05
            qty_0b1 = self.num(row_0b1.get("數量", ""))
            row_0b1["來源編號"] = "01"
            row_0b1["比例"] = "0.05"
            row_0b1["單價"] = self.fmt_2(tax_unit_price)
            row_0b1["複價"] = self.fmt_2(tax_unit_price * qty_0b1) if qty_0b1 is not None else self.fmt_2(tax_unit_price)
            changed_cells.extend([("0B1", "來源編號"), ("0B1", "比例"), ("0B1", "單價"), ("0B1", "複價")])
            changed += 1

        if row_0b is not None:
            child_amount = self.num(row_0b1.get("複價", "")) if row_0b1 is not None else None
            if child_amount is None and base_01 is not None:
                child_amount = base_01 * 0.05
            if child_amount is not None:
                qty_0b = self.num(row_0b.get("數量", ""))
                row_0b["單價"] = self.fmt_2(child_amount)
                row_0b["複價"] = self.fmt_2(child_amount * qty_0b) if qty_0b is not None else self.fmt_2(child_amount)
                changed_cells.extend([("0B", "單價"), ("0B", "複價")])
                changed += 1

        return changed

    def calculate_rollup_amounts(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        self.commit_active_editor()
        self.sync_table_widget_to_data(self.combined_table, self.combined_data)
        self.push_edit_history()

        item_to_row = self.item_map()
        items = sorted(item_to_row.keys(), key=len, reverse=True)

        leaf_changed = 0
        rollup_changed = 0
        dynamic_changed = 0
        skipped = 0
        changed_cells = []

        for item in items:
            if item.upper() in {"0B", "0B1"}:
                continue
            if self.rollup_children(item, item_to_row) and item.upper() not in RIGHT_TOP_0A_CODES:
                continue

            row = item_to_row[item]
            qty = self.num(row.get("數量", ""))
            price = self.num(row.get("單價", ""))

            if qty is not None and price is not None:
                row["複價"] = self.fmt_2(qty * price)
                changed_cells.append((item, "複價"))
                leaf_changed += 1
            else:
                skipped += 1

        dynamic_changed, dynamic_cells = self.sync_dynamic_calculations_to_combined(item_to_row)
        changed_cells.extend(dynamic_cells)
        rollup_changed, rollup_skipped = self.rollup_parent_amounts(item_to_row, changed_cells)
        skipped += rollup_skipped

        # 先完成 01 到 0A 的工項計算，再用 01 複價 × 5% 回填 0B1，並重算 0B。
        tax_changed = self.calculate_0b1_and_0b(item_to_row, changed_cells)
        dynamic_changed += tax_changed
        total_project_changed = self.calculate_total_project_from_second_level(item_to_row, changed_cells)
        rollup_changed += total_project_changed

        self.recalculated_amounts = self.display_amount_snapshot(calculated_side=True)
        self.refresh_combined_and_final(self.combined_data)
        self.mark_calc2_blue(changed_cells)

        self.msg_info(
            "完成",
            f"重新計算完成\n"
            f"末階計算：{leaf_changed} 筆\n"
            f"階層彙總：{rollup_changed} 筆\n"
            f"指定工項複價計算：{dynamic_changed} 筆\n"
            f"略過：{skipped} 筆"
        )

    # =========================================================
    # 刪除單複價規則
    # =========================================================
    def apply_delete_rules(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        self.push_edit_history()

        unit_items = {
            str(row.get("項目", "")).strip()
            for row in self.unit_price_data
            if str(row.get("項目", "")).strip()
        }

        combined_items = [
            str(row.get("項目", "")).strip()
            for row in self.combined_data
            if str(row.get("項目", "")).strip()
        ]

        def has_child(item):
            return any(
                other != item and other.startswith(item) and len(other) > len(item)
                for other in combined_items
            )

        def has_unit_child(item):
            return any(
                other != item and other.startswith(item) and len(other) > len(item)
                for other in unit_items
            )

        def delete_price_allowed(row, item):
            if has_child(item) and has_unit_child(item):
                row["單價"] = ""

        changed = 0

        for row in self.combined_data:
            item = str(row.get("項目", "")).strip()

            if not item:
                continue

            old = (row.get("單價", ""), row.get("複價", ""))
            prefix2 = item[:2].upper()
            item_len = len(item)
            is_unit_layer = item in unit_items

            if item == "0":
                row["單價"] = ""
                row["複價"] = ""

            elif item.startswith("011"):
                if 2 <= item_len <= 5:
                    delete_price_allowed(row, item)
                    row["複價"] = ""
                elif item_len == 6:
                    row["複價"] = ""

            elif item.startswith("012"):
                row["複價"] = ""

            elif item.startswith("013"):
                if item in {"013", "0131", "0132", "0133"}:
                    delete_price_allowed(row, item)
                elif has_unit_child(item):
                    row["單價"] = ""
                row["複價"] = ""

            elif item.startswith("014"):
                if item in {"014", "0141", "0142"}:
                    row["單價"] = ""
                    row["複價"] = ""
                elif is_unit_layer:
                    row["複價"] = ""
                elif has_unit_child(item):
                    row["單價"] = ""
                    row["複價"] = ""
                else:
                    row["複價"] = ""

            elif item == "01" or item.startswith("01"):
                delete_price_allowed(row, item)
                row["複價"] = ""

            elif item.upper() in {"0B", "0B1"}:
                row["單價"] = ""
                row["複價"] = ""
                if item.upper() == "0B1":
                    row["來源編號"] = "01"
                    row["比例"] = "0.05"

            elif self.is_02_to_0a(item):
                if item_len == 2:
                    if prefix2 in {"02", "0A"}:
                        row["單價"] = ""
                    elif prefix2 not in {"03", "04", "05", "06"}:
                        delete_price_allowed(row, item)
                    row["複價"] = ""
                else:
                    row["複價"] = ""

            if old != (row.get("單價", ""), row.get("複價", "")):
                changed += 1

        self.refresh_combined_and_final(self.combined_data)
        self.mark_remaining_price_amount_red()
        self.msg_info("完成", f"刪除單複價完成\n影響 {changed} 筆資料")

    def is_02_to_0a(self, item):
        return len(item) >= 2 and item[:2].upper() in {
            "02", "03", "04", "05", "06", "07", "08", "09", "0A"
        }

    # =========================================================
    # 整理備註
    # =========================================================
    def clean_remark_string(self, remark):
        text = "" if remark is None else str(remark)

        if text.strip() == "":
            return ""

        # V4.3.3 整理備註規則：
        # 從左側開始，若前段字串後接「,#」、「,*」或「,#,*」，
        # 刪除該符號左側與分隔符本身，只保留後方字串。
        candidates = []
        for marker in (",#,*", ",#", ",*"):
            index = text.find(marker)
            if index >= 0:
                candidates.append((index, marker))

        if not candidates:
            return text

        index, marker = min(candidates, key=lambda item: item[0])
        return text[index + len(marker):]

    def apply_clean_remark_rules(self):
        changed = 0

        for row in self.combined_data:
            old = row.get("備註", "")
            new = self.clean_remark_string(old)

            if old != new:
                row["備註"] = new
                changed += 1

        return changed

    def apply_clean_remark2_rules(self):
        deleted_rows = 0
        deleted_words = 0
        matched_rows = 0
        matched_words = 0

        for row in self.combined_data:
            remark = "" if row.get("備註", "") is None else str(row.get("備註", ""))

            while True:
                match = REMARK2_LEADING_TOKEN_PATTERN.match(remark)

                if not match:
                    break

                deleted_words += 1
                remark = remark[match.end():]
                separator = REMARK2_LEADING_SEPARATOR_PATTERN.match(remark)

                if separator:
                    remark = remark[separator.end():]
                else:
                    remark = remark.lstrip()

            if remark != ("" if row.get("備註", "") is None else str(row.get("備註", ""))):
                row["備註"] = remark
                deleted_rows += 1

            matches = REMARK2_PATTERN.findall(remark)

            if matches:
                matched_rows += 1
                matched_words += len(matches)

        return deleted_rows, deleted_words, matched_rows, matched_words

    def organize_combined_remarks(self):
        if not self.combined_data:
            self.msg_warn("提醒", "組合總表尚無資料，請先分析。")
            return

        self.push_edit_history()

        clean_changed = self.apply_clean_remark_rules()
        deleted_rows, deleted_words, matched_rows, matched_words = self.apply_clean_remark2_rules()
        self.remark2_highlight_enabled = True
        self.refresh_combined_and_final(self.combined_data)
        self.combined_table.viewport().update()
        self.status_label.setText(
            f"備註整理完成｜分隔符整理 {clean_changed} 列｜左側刪除 {deleted_rows} 列/{deleted_words} 個｜藍字 {matched_rows} 列/{matched_words} 個"
        )
        self.msg_info(
            "完成",
            f"備註整理完成\n"
            f"檢查欄位：第四分頁備註\n"
            f"分隔符整理：{clean_changed} 列\n"
            f"檢查規則：英文字母開頭且含數字，或數字開頭且含英文字母\n"
            f"左側刪除：{deleted_words} 個字串，影響 {deleted_rows} 列\n"
            f"剩餘藍字標示：{matched_words} 個字串，影響 {matched_rows} 列"
        )

    # =========================================================
    # 輸出總表動態填入固定欄位
    # =========================================================
    def fill_final_fixed_fields(self):
        """
        將目前輸入欄位：
        - 檔別
        - 動支單號
        - 單位檢核末碼

        動態填入輸出總表所有資料列。
        """
        if not self.final_data:
            self.msg_warn("提醒", "輸出總表目前沒有資料可填入。")
            return

        self.push_edit_history()

        file_type = self.file_type_edit.text()
        payment_no = self.payment_no_edit.text()
        unit_check = self.unit_check_edit.text()

        for row in self.final_data:
            row["檔別"] = file_type
            row["動支單號"] = payment_no
            row["單位檢核末碼"] = unit_check

        self.populate_table(self.final_table, self.final_data)
        self.status_label.setText(
            f"已將檔別、動支單號、單位檢核末碼填入輸出總表｜共 {len(self.final_data)} 筆"
        )
        self.save_edit_log(show_message=False)
        self.msg_info("完成", f"已填入輸出總表固定欄位\n共 {len(self.final_data)} 筆")

    # =========================================================
    # 匯入回饋 Excel
    # =========================================================
    def normalize_import_rows(self, raw_rows):
        if not raw_rows:
            return []

        headers = [str(h).strip() for h in raw_rows[0]]
        index_map = {header: i for i, header in enumerate(headers)}
        rows = []

        for raw in raw_rows[1:]:
            if not any(str(value).strip() for value in raw):
                continue

            row = {}

            for col in COLUMNS:
                idx = index_map.get(col)
                row[col] = "" if idx is None or idx >= len(raw) else str(raw[idx]).strip()

            rows.append(row)

        return rows

    def read_html_xls_rows(self, filename):
        content = None

        for enc in ("utf-8-sig", "utf-8", "cp950", "big5"):
            try:
                content = Path(filename).read_text(encoding=enc)
                break
            except Exception:
                pass

        if content is None:
            raise ValueError("無法讀取 XLS 檔案。若這是舊版二進位 .xls，請先另存成 .xlsx 或用本程式匯出的 .xls。")

        parser = SimpleHTMLTableParser()
        parser.feed(content)

        return self.normalize_import_rows(parser.rows)

    def read_openpyxl_rows(self, filename):
        load_workbook = ensure_package("openpyxl", "openpyxl").load_workbook
        wb = load_workbook(filename, data_only=True, read_only=True)
        ws = wb.active
        raw_rows = []

        for row in ws.iter_rows(values_only=True):
            raw_rows.append(["" if cell is None else str(cell) for cell in row])

        return self.normalize_import_rows(raw_rows)

    def import_feedback_excel(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "匯入回饋 Excel", "",
            "Excel (*.xls *.xlsx *.xlsm);;All Files (*.*)"
        )

        if not filename:
            return

        try:
            lower = filename.lower()

            if lower.endswith(".xls"):
                rows = self.read_html_xls_rows(filename)
            elif lower.endswith(".xlsx") or lower.endswith(".xlsm"):
                rows = self.read_openpyxl_rows(filename)
            else:
                self.msg_error("錯誤", "只支援 .xls、.xlsx、.xlsm")
                return

            if not rows:
                self.msg_warn("提醒", "匯入檔案沒有可用資料，請確認第一列是欄位名稱。")
                return

            clean_rows = []

            for row in rows:
                new_row = {col: row.get(col, "") for col in COLUMNS}

                for col in ("數量", "單價", "複價"):
                    new_row[col] = self.final_number_text(new_row.get(col, ""))

                clean_rows.append(new_row)

            self.push_edit_history()

            self.final_data = clean_rows
            self.combined_data = [dict(row) for row in clean_rows]

            self.populate_table(self.combined_table, self.combined_data)
            self.populate_table(self.final_table, self.final_data)

            self.status_label.setText(f"已匯入回饋 Excel：{len(clean_rows)} 筆")
            self.save_edit_log(show_message=False)
            self.msg_info("完成", f"已匯入回饋 Excel：{len(clean_rows)} 筆")

        except Exception:
            self.msg_error("錯誤", traceback.format_exc())

    # =========================================================
    # 匯出 Excel
    # =========================================================
    def get_final_export_rows(self):
        if not self.final_data and self.combined_data:
            self.final_data = self.make_final_rows(self.combined_data)
            self.populate_table(self.final_table, self.final_data)

        return self.final_data

    def safe_filename_part(self, text, fallback):
        value = "" if text is None else str(text).strip()
        value = re.sub(r'[\\/:*?"<>|]+', "_", value)
        value = re.sub(r"\s+", "", value)
        value = value.strip("._")
        return value or fallback

    def default_export_filename_english(self, extension):
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        project_no = self.safe_filename_part(self.project_no_edit.text(), "未填工程編號")
        return str(Path.cwd() / f"AA3118_{project_no}_{timestamp}.{extension}")

    def default_export_filename(self, extension):
        return self.default_export_filename_english(extension)

    def export_final_xlsx(self):
        rows = self.get_final_export_rows()

        if not rows:
            self.msg_warn("提醒", "輸出總表沒有資料可匯出。")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "匯出 XLSX", self.default_export_filename("xlsx"), "Excel Workbook (*.xlsx)"
        )

        if not filename:
            return

        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"

        try:
            openpyxl = ensure_package("openpyxl", "openpyxl")
            Workbook = openpyxl.Workbook
            from openpyxl.styles import Font, Alignment
            wb = Workbook()
            wb.properties.creator = "預算書分析系統"
            wb.properties.title = "輸出總表"
            wb.properties.subject = "UTF-8 繁體中文輸出"

            ws = wb.active
            ws.title = "輸出總表"

            for c, col in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=1, column=c)
                cell.value = str(col)
                cell.data_type = "s"
                cell.number_format = "@"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for r, row in enumerate(rows, start=2):
                for c, col in enumerate(COLUMNS, start=1):
                    cell = ws.cell(row=r, column=c)
                    cell.value = "" if row.get(col, "") is None else str(row.get(col, ""))
                    cell.data_type = "s"
                    cell.number_format = "@"

            for c, col in enumerate(COLUMNS, start=1):
                max_len = len(str(col))
                for row in rows[:1000]:
                    max_len = max(max_len, len(str(row.get(col, ""))))
                ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(max(max_len + 2, 8), 40)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            wb.save(filename)
            self.msg_info("完成", f"XLSX 匯出完成：\n{filename}")

        except Exception:
            self.msg_error("錯誤", traceback.format_exc())


def main():
    app = QApplication(sys.argv)
    win = BudgetAnalyzerQT()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()


